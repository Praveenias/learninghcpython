#####################################################
#  Dependency: openpyxl, argparse
#  It shows the example of split, as well as processing individual column
#####################################################

import os 
import sys
import pathlib 
import math 
import re
import pprint
import argparse

## To Read Excel work book
from openpyxl import load_workbook

## To write Excel work book
from openpyxl import Workbook


suffix_l = ['inc', 'corp', 'corporation', 'llc', 'ltd', 'limited']

def convert_string(raw_string):
  '''
    Trim the string
    Convert all non-Alphanumeric characters to _ 
    Convert to lowercase
    If the last word is any one of suffix_l then remove that ( useful
     to normalize manufacturer name )
    Reduce consequtive _ characters into single _
  '''
  if (not raw_string):
    return None
  tmp_s = str(raw_string).strip()
  tmp_s = re.sub('[^\w]+', ' ',tmp_s)
  tmp_l = tmp_s.lower().split()
  if (tmp_l[-1] in suffix_l): del tmp_l[-1]
  res = '_'.join(tmp_l)
  return res 


def row2dictionary(row, header_l):
  '''
    Convert given row to dictionary for all the fields in header_l and
    also in all_fd_l 
    Note: This converts all field value to string and strip them
    It addes norm_mfr and norm_mpn fields 
  '''
  res_d = { }

  for idx, head in enumerate(header_l):
    #~ Handle hyperlink fields
    if (head == 'data_sheet_document'):
      try:
        res_d[head] = row[idx].hyperlink.target.strip()
      except AttributeError as attr_err:  # No hyperlink present
        res_d[head] = row[idx].value
    else:
      res_d[head] = str(row[idx].value).strip()
  
  #~ Add these two fields
  res_d['norm_mfr'] = convert_string(res_d['manufacturer_name'])
  res_d['norm_mpn'] = convert_string(res_d['manufacturer_part_number'])
  return res_d

def write2sheet(ip_sheet, op_fname, op_sheetname, st_row, e_row, header_l=None):

  result = {
      'status':'Success',
      'msg':'',
      }
  try:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = op_sheetname
    if (header_l):
      ws1.append(header_l)
    
    for row in ip_sheet.iter_rows(min_row = st_row, max_row = e_row):
      row_l = [ cell.value for cell in row]
      ws1.append(row_l)
    wb.save(op_fname)
  except Exception as err:
    result['status'] = 'Failure'
    result['msg'] = err
  return result


def split_sheet(ip_fname, op_basename, rows_per_file=0, no_of_files=0,
                ip_sheetname='Data', op_sheetname='Data', hdr_row = 1,
                dry_run=False):

  result = {
    'status':'Success',
    'msg'   : '',
    }

  ip_path = pathlib.Path(ip_fname) 
  if ( not ip_path.is_file()):
    result['status'] = 'Failure'
    result['msg'] = (f'Input is not a valid file: {ip_fname}')
    return result

  if (hdr_row < 1):
    result['status'] = 'Failure'
    result['msg'] = (f'Invalid header row number: {hdr_row}')
    return result

  ## Eiter one of the rows_per_file , no_of_files should be given
  if ( (rows_per_file < 1) and (no_of_files < 1) ):
    result['status'] = 'Failure'
    tmp ='Both rows_per_file and no_of_files cannot be less than 1: ' 
    result['msg'] = (f'{tmp} {rows_per_file}, {no_of_files}')
    return result

  try: 
    #~ If we set read only then hyper link is not read
    ip_workbook = load_workbook(filename = ip_fname, data_only=True)
    if (ip_sheetname):
      ip_sheet = ip_workbook[ip_sheetname]
    else:
      ip_sheet = ip_workbook.active

    #~ Convert header row to List
    header_l = [ cell.value for cell in ip_sheet[hdr_row] ] 
  except Exception as err:
    result['status'] = 'Failure'
    result['msg']    = err
    return result

  if (no_of_files < 1):
    no_of_files = math.ceil(ip_sheet.max_row / rows_per_file)
  else:
    rows_per_file = math.ceil(ip_sheet.max_row / no_of_files)

  for i in range(1,no_of_files+1):
    suffix = f'{i:03d}'
    op_fname = op_basename+'_'+suffix+'.xlsx'
    if (i==1):
      st_row = hdr_row+1 
    else:
      st_row = e_row
    e_row  = st_row+rows_per_file+1

    #~ Create output files only if it is not dry_run
    if (not dry_run):
      wr_result = write2sheet(ip_sheet, op_fname, op_sheetname, st_row, e_row, 
                              header_l)
      if (wr_result['status'] != 'Success'):
        result['status'] = 'Failure'
        result['msg'] = f'Writing {i}th file failed. {wr_result["msg"]}'
        break

  if (result['status'] == 'Success'):
    if (not dry_run):
      result['msg'] = f'{no_of_files} created with base name: {op_basename}'
    else:
      result['msg']  = f'{no_of_files} will be created with base name: '
      result['msg'] += f'{op_basename}'

  return result

def process_sheet(ip_fname, ip_sheetname='Data', header_row = 1, min_row = None,
                  no_rows = None):

  result = {
    'status':'Success',
    'msg'   : '',
    }

  if ( (min_row) and (min_row > 0) ):
    min_row = min_row
  else:
    min_row = header_row + 1
  if ( (no_rows) and (no_rows > 0) ):
    max_row = min_row + no_rows - 1 
  else:
    max_row = None

  try: 
    #~ If we set read only then hyper link is not read
    ip_workbook = load_workbook(filename = ip_fname, data_only=True)
    if (ip_sheetname):
      ip_sheet = ip_workbook[ip_sheetname]
    else:
      ip_sheet = ip_workbook.active

    #~ Convert header row to List
    header_l = [ cell.value for cell in ip_sheet[hdr_row] ] 
  except Exception as err:
    result['status'] = 'Failure'
    result['msg']    = err
    return result

  header_l = [ convert_string(cell.value) for cell in ip_sheet[header_row]]
  for row in ip_sheet.iter_rows(min_row = min_row, max_row = max_row):
    row_d = row2dictionary(row,header_l)
    pp.pprint(row_d)
  return result



#####################################################
#  Start of Main Program
#####################################################


print('Start of  Excel split')
pp = pprint.PrettyPrinter(indent=2)
'''
## Command Line Processing
cli_parser = argparse.ArgumentParser(
    prog='split_excel',
    description='Split given work sheet into multiple files')

# Add the arguments
cli_parser.add_argument('--input', '-i', type=str, required=True,
                        help='Input Excel file')

cli_parser.add_argument('--sheet', '-s', type=str, required=False,
                help="Input and Output Sheet Name", default='Data')

cli_parser.add_argument('--header', '-r', type=int, required=False,
                        help='Header Row Number', default=2)

cli_parser.add_argument('--output', '-o', type=str, required=True,
                        help='Output Excel file Base Name')

cli_parser.add_argument('--norows', '-n', type=int, required=False,
               help='Number of Rows.If no of files given this will be ignored',
                        default=0)

cli_parser.add_argument('--nofiles', '-f', type=int, required=False,
                        help='Number of Files', default=0)

cli_parser.add_argument('--opt', '-p', type=str, required=False,
                        help='Command line option File')

cli_parser.add_argument('--dry', '-d', type=bool, required=False,
                        help='Whether Dry run or not', default=False)

args = cli_parser.parse_args()
#~ pp.pprint(args)

ip_fname       = args.input
ip_sheetname   = args.sheet
op_basename    = args.output
op_sheetname   = args.sheet
rows_per_file  = args.norows
no_of_files    = args.nofiles
dry_run        = args.dry
'''

ip_fname       = '../data/mastercomp_25.xlsx'
ip_sheetname   = 'Data'
op_basename    = '../data/result'
op_sheetname   = 'Data'
rows_per_file  = 5
no_of_files    = 0
dry_run        = True


#~ print(ip_fname, ip_sheetname, op_basename, op_sheetname, rows_per_file,
         #~ no_of_files, dry_run)
'''
result = split_sheet(ip_fname, op_basename, rows_per_file=rows_per_file, 
                     no_of_files=no_of_files,
                     ip_sheetname=ip_sheetname, op_sheetname=op_sheetname,
                     dry_run=dry_run)
'''
hdr_row = 1
result = process_sheet(ip_fname, ip_sheetname, min_row=2,no_rows=1) 
pp.pprint(result)


'''
ip_path = pathlib.Path(ip_fname) 
if ( not ip_path.is_file()):
  print(f'Input is not a valid file: {ip_fname}')

#~ If we set read only then hyper link is not read
ip_workbook = load_workbook(filename = ip_fname, data_only=True)
print('workbook loaded')


if (ip_sheetname):
  in_sheet = ip_workbook[ip_sheetname]
else:
  in_sheet = ip_workbook.active
  
#~ Convert header row to List
header_l = [ cell.value for cell in in_sheet[1] ] 

#~ Some information for debug
print('Input Header Row: ', header_l)
print('Input Number of Rows: ', in_sheet.max_row)
print('Input Number of columns: ', in_sheet.max_column)
'''


