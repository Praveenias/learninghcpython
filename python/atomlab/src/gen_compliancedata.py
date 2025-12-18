"""
  Given any excel file contain mastercomponents, it extracts unique
  MFR and MPN from that file and then create compliance data file as required
  by the ATOM template
"""
import pprint
from datetime import datetime
from pathlib import Path
import argparse
from argparse import Namespace
import re
from typing import MutableMapping
from faker import Faker
import random

import openpyxl as xl

## To write Excel work book
from openpyxl.workbook import workbook

column_names = [
    "MFR Name",
    "MPN",
    "RoHS Status",
    "RoHS Version",
    "Exemption Annex",
    "RoHS exemptions",
    "RoHS Source Doc Name",
    "RoHS Source Doc",
    "RoHS Doc Date",
    "RoHS Doc Type",
    "RoHS Doc Link",
    "REACH SVHC Status",
    "REACH SVHC Version",
    "REACH SVHC Source Doc Name",
    "REACH SVHC Source Doc",
    "REACH SVHC Doc Date",
    "REACH SVHC Doc Type",
    "REACH SVHC Doc Link",
    "REACH Annex XIV Version",
    "REACH Annex XIV Status",
    "REACH Annex XIV Source Doc Name",
    "REACH Annex XIV Source Doc ",
    "REACH Annex XIV Doc Date",
    "REACH Annex XIV Doc Type",
    "REACH Annex XIV Doc Link",
    "REACH Annex XVII Version",
    "REACH Annex XVII Status",
    "REACH Annex XVII  Source Doc",
    "REACH Annex XVII Source Doc Name",
    "REACH Annex XVII Doc Date",
    "REACH Annex XVII Doc Type",
    "REACH Annex XVII Doc Link",
    "CA Pro 65 Version",
    "CA Pro 65 Status",
    "CA Pro 65 Potential Exposure",
    "CA Pro 65 Warning",
    "CA Pro 65 Source Doc Name",
    "CA Pro 65 Source Doc",
    "CA Pro 65 Doc Date",
    "CA Pro 65 Doc Type",
    "CA Pro 65 Doc Link",
    "TSCA Version",
    "TSCA Status",
    "TSCA Source Doc Name",
    "TSCA Source Doc",
    "TSCA Doc Date",
    "TSCA Doc Type",
    "TSCA Doc Link",
]

suffix_l = ["inc", "corp", "corporation", "llc", "ltd", "limited"]


def normalize_string(raw_string):
    """
    Trim the string
    Convert all non-Alphanumeric characters to _
    Convert to lowercase
    If the last word is any one of suffix_l then remove that ( useful
     to normalize manufacturer name )
    Reduce consequtive _ characters into single _
    """
    if not raw_string:
        return None
    tmp_s = str(raw_string).strip()
    tmp_s = re.sub("[^\w]+", " ", tmp_s)
    tmp_l = tmp_s.lower().split()
    if tmp_l[-1] in suffix_l:
        del tmp_l[-1]
    res = "_".join(tmp_l)
    return res


def find_mfr_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("manufacturername")
    except ValueError:
        column_number = -1
    return column_number


def find_mpn_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("manufacturerpartnumber")
    except ValueError:
        column_number = -1
    return column_number


def get_mfr_names(ip_sheet: xl.worksheet, hdr_row_no: int, mfr_col: int) -> list:
    st_row = hdr_row_no + 1
    mfr_names = [
        ip_sheet[r][mfr_col].value for r in range(st_row, ip_sheet.max_row + 1)
    ]
    return mfr_names


def get_mpn_names(ip_sheet: xl.worksheet, hdr_row_no: int, mpn_col: int) -> list:
    st_row = hdr_row_no + 1
    mpn_names = [
        ip_sheet[r][mpn_col].value for r in range(st_row, ip_sheet.max_row + 1)
    ]
    return mpn_names


def get_mfr_record(mfr: str, mpn: str) -> list:
    """
    ToDo: Need to fill up all the mandatory column
    """
    RoHS_Status = ["Compliant","Compliant with exemption"]
    RoHS_Status = random.choice(RoHS_Status)
    RoHS_Version = ["EUROHS-1907", "EUROHS-1506", "EUROHS-0509"]
    RoHS_Version = random.sample(RoHS_Version, 1)
    Exemption_Annex = ["Annex IV", " Annex III"]
    Exemption_Annex = random.sample(Exemption_Annex, 1)
    RoHs_exemptions = [None, "6c", "18b1", "2a1", "2b1"]
    RoHs_exemptions = random.sample(RoHs_exemptions, 1)
    RoHs_Source_Doc = ["Test Report", "coC"]
    RoHs_Source_Doc = random.choice(RoHs_Source_Doc)
    RoHs_Source_Doc_Name = ["coC", "FMD"]
    RoHs_Source_Doc_Name = random.sample(RoHs_Source_Doc_Name, 1)
    RoHs_Doc_Date = ["2020-02-25"]
    RoHs_Doc_Type = ["Partseries", "Blanket"]
    RoHs_Doc_Type = random.sample(RoHs_Doc_Type, 1)
    RoHS_Doc_Link = [
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N3428-5203RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N2514-6002-RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3385-6614_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4-CUST_3M_3421-6600_AND_3448-3020_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_8209-6000_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3448-8D09A_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3473-6610_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_89126-0103_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3421-6620_072420.pdf",
    ]
    RoHS_Doc_Link = random.choice(RoHS_Doc_Link)
    REACH_SVHC_Status = [
        "Non Compliant",
        "Compliant",
    ]
    REACH_SVHC_Status = random.choice(REACH_SVHC_Status)
    REACH_SVHC_Version = [
        "EUREACH-0620",
        "EUREACH-0721",
        "EUREACH-0119",
        "EUREACH-0120",
    ]
    REACH_SVHC_Version = random.choice(REACH_SVHC_Version)
    REACH_SVHC_Source_Doc = ["Test Report", "CoC"]
    REACH_SVHC_Source_Doc = random.choice(REACH_SVHC_Source_Doc)
    REACH_SVHC_Source_Doc_Date = ["2020-02-25"]
    REACH_SVHC_Source_Doc_Date = random.sample(REACH_SVHC_Source_Doc_Date, 1)
    REACH_SVHC_Source_Doc_Name = ["COC"]
    REACH_SVHC_Source_Doc_Type = ["Partseries", "Blanket"]
    REACH_SVHC_Source_Doc_Type = random.sample(REACH_SVHC_Source_Doc_Type, 1)
    REACH_SVHC_Source_Doc_Link = [
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N3428-5203RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N2514-6002-RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3385-6614_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4-CUST_3M_3421-6600_AND_3448-3020_072420.pdf",
    ]
    REACH_SVHC_Source_Doc_Link = random.sample(REACH_SVHC_Source_Doc_Link, 1)
    REACH_Annex_XIV_Version = ["REACH Annex XIV"]
    REACH_Annex_XIV_Version = random.sample(REACH_Annex_XIV_Version, 1)
    REACH_Annex_XIV_Status = [
        "Compliant",
        "Non Compliant",
    ]
    REACH_Annex_XIV_Status = random.sample(REACH_Annex_XIV_Status, 1)
    REACH_Annex_XIV_Source_Doc = ["Test Report", "CoC"]
    REACH_Annex_XIV_Source_Doc = random.sample(REACH_Annex_XIV_Source_Doc, 1)
    REACH_Annex_XIV_Doc_Date = ["2020-02-25"]
    REACH_Annex_XIV_Doc_Type = ["Partseries", "Blanket"]
    REACH_Annex_XIV_Doc_Type = random.sample(REACH_Annex_XIV_Doc_Type, 1)
    REACH_Annex_XIV_Source_Doc_Name = ["COC"]
    REACH_Annex_XIV_Doc_Link = [
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N3428-5203RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N2514-6002-RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3385-6614_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4-CUST_3M_3421-6600_AND_3448-3020_072420.pdf",
    ]
    REACH_Annex_XIV_Doc_Link = random.sample(REACH_Annex_XIV_Doc_Link, 1)
    REACH_Annex_XVII_VERSION = [" REACH Annex XVII"]
    REACH_Annex_XVII_VERSION = random.sample(REACH_Annex_XVII_VERSION, 1)
    REACH_Annex_XVII_STATUS = ["Non Compliant", "Compliant"]
    REACH_Annex_XVII_STATUS = random.sample(REACH_Annex_XVII_STATUS, 1)
    REACH_Annex_XVII_Source_Doc = ["Test Report", "coC"]
    REACH_Annex_XVII_Source_Doc = random.sample(REACH_Annex_XVII_Source_Doc, 1)
    REACH_Annex_XVII_Source_Doc_Name = ["COC"]
    REACH_Annex_XVII_Doc_Date = ["2020-02-25"]
    REACH_Annex_XVII_Doc_Type = ["Partseries", "Blanket"]
    REACH_Annex_XVII_Doc_Type = random.sample(REACH_Annex_XVII_Doc_Type, 1)
    REACH_Annex_XVII_Doc_Link = [
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N3428-5203RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N2514-6002-RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4-CUST_3M_3421-6600_AND_3448-3020_072420.pdf",
    ]
    REACH_Annex_XVII_Doc_Link = random.sample(REACH_Annex_XVII_Doc_Link, 1)
    CA_PRO_65_Version = ["CA Prop 65-0120"]
    CA_PRO_65_Status = ["Affected", "Not Applicable", "Not Affected"]
    CA_PRO_65_Status = random.sample(CA_PRO_65_Status, 1)
    CA_PRO_65_Potential_Exposure = ["Yes", "No", "No Data","Not Applicable"]
    CA_PRO_65_Potential_Exposure = random.sample(CA_PRO_65_Potential_Exposure, 1)
    CA_PRO_65_Warning = ["Available", "No Data", "Not Applicable"]
    CA_PRO_65_Warning = random.sample(CA_PRO_65_Warning, 1)
    CA_PRO_65_Source_Doc_Name = ["COC-1", "COC-2", "COC-3", "COC-4", "COC-5", "COC-6"]
    CA_PRO_65_Source_Doc_Name = random.sample(CA_PRO_65_Source_Doc_Name, 1)
    CA_PRO_65_Source_Doc = ["Test Report", "coC"]
    CA_PRO_65_Source_Doc = random.sample(CA_PRO_65_Source_Doc, 1)
    CA_PRO_65_Doc_Date = ["2020-02-25"]
    CA_PRO_65_Doc_Type = ["Partseries", "Blanket"]
    CA_PRO_65_Doc_Type = random.sample(CA_PRO_65_Doc_Type, 1)
    CA_PRO_65_Doc_Link = [
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N3428-5203RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N2514-6002-RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3385-6614_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4-CUST_3M_3421-6600_AND_3448-3020_072420.pdf",
    ]
    CA_PRO_65_Doc_Link = random.sample(CA_PRO_65_Doc_Link, 1)
    TSCA_Version = ["TSCA-PBT"]
    TSCA_Status = ["Compliant", "Non Compliant"]
    TSCA_Status = random.sample(TSCA_Status, 1)
    TSCA_Source_Doc_Name = ["COC-1", "COC-2", "COC-3", "COC-4", "COC-5", "COC-6"]
    TSCA_Source_Doc_Name = random.sample(TSCA_Source_Doc_Name, 1)
    TSCA_Source_Doc = ["Test Report", "coC"]
    TSCA_Source_Doc = random.sample(TSCA_Source_Doc, 1)
    TSCA_Doc_Date = ["2020-02-25"]
    TSCA_Doc_Type = ["Partseries", "Blanket"]
    TSCA_Doc_Type = random.sample(TSCA_Doc_Type, 1)
    TSCA_Doc_Link = [
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N3428-5203RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_N2514-6002-RB_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4_3M_3385-6614_072420.pdf",
        "https://www.compliance.ilensys.com/compliance_docs/CoC4-CUST_3M_3421-6600_AND_3448-3020_072420.pdf",
    ]
    TSCA_Doc_Link = random.sample(TSCA_Doc_Link, 1)

    return [
        mfr,
        mpn,
        RoHS_Status,
        *RoHS_Version,
        *Exemption_Annex,
        *RoHs_exemptions,
        *RoHs_Source_Doc_Name,
        RoHs_Source_Doc,
        *RoHs_Doc_Date,
        *RoHs_Doc_Type,
        RoHS_Doc_Link,
        REACH_SVHC_Status,
        REACH_SVHC_Version,
        *REACH_SVHC_Source_Doc_Name,
        REACH_SVHC_Source_Doc,
        *REACH_SVHC_Source_Doc_Date,
        *REACH_SVHC_Source_Doc_Type,
        *REACH_SVHC_Source_Doc_Link,
        *REACH_Annex_XIV_Version,
        *REACH_Annex_XIV_Status,
        *REACH_Annex_XIV_Source_Doc_Name,
        *REACH_Annex_XIV_Source_Doc,
        *REACH_Annex_XIV_Doc_Date,
        *REACH_Annex_XIV_Doc_Type,
        *REACH_Annex_XIV_Doc_Link,
        *REACH_Annex_XVII_VERSION,
        *REACH_Annex_XVII_STATUS,
        *REACH_Annex_XVII_Source_Doc,
        *REACH_Annex_XVII_Source_Doc_Name,
        *REACH_Annex_XVII_Doc_Date,
        *REACH_Annex_XVII_Doc_Type,
        *REACH_Annex_XVII_Doc_Link,
        *CA_PRO_65_Version,
        *CA_PRO_65_Status,
        *CA_PRO_65_Potential_Exposure,
        *CA_PRO_65_Warning,
        *CA_PRO_65_Source_Doc_Name,
        *CA_PRO_65_Source_Doc,
        *CA_PRO_65_Doc_Date,
        *CA_PRO_65_Doc_Type,
        *CA_PRO_65_Doc_Link,
        *TSCA_Version,
        *TSCA_Status,
        *TSCA_Source_Doc_Name,
        *TSCA_Source_Doc,
        *TSCA_Doc_Date,
        *TSCA_Doc_Type,
        *TSCA_Doc_Link,
    ]


def create_manufacturer(
    args: Namespace,
    output_fname: str,
) -> dict:
    """
    Create manufacturer recrods for all the unique manufacturer given
    in master component file
    Args:
      args(dictionary):
        input : Input master component file name ( with path )
        sheet : Sheet name , containing input data ( if not given first sheet )
        header : Header row number. Data assumed to one row after header till
                  end of the file
    Returns:
     Dictionay
      status: 'Success' or 'Failure'
      msg:     If failure will contain error message else empty string
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    try:
        # ~ If we set read only then hyper link is not read
        ip_workbook = xl.load_workbook(filename=args.input, data_only=True)
        if args.sheet:
            ip_sheet = ip_workbook[args.sheet]
        else:
            ip_sheet = ip_workbook.active
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result

    mfr_column = find_mfr_column(ip_sheet, args.header)
    if mfr_column < 0:
        result["status"] = "Failure"
        result["msg"] = "Manufacturer Name column not present in input file"
        return result

    mpn_column = find_mpn_column(ip_sheet, args.header)
    if mpn_column < 0:
        result["status"] = "Failure"
        result["msg"] = "Manufacturer part number column not present in input file"
        return result

    mfr_names = get_mfr_names(ip_sheet, args.header, mfr_column)
    mpn_names = get_mpn_names(ip_sheet, args.header, mpn_column)

    master_workbook = xl.Workbook()
    master_worksheet = master_workbook.active
    master_worksheet.title = "Data"

    try:
        ## First append header row
        master_worksheet.append(column_names)
        for mfr, mpn in zip(mfr_names, mpn_names):
            mfr_record = get_mfr_record(mfr, mpn)
            master_worksheet.append(mfr_record)
        master_workbook.save(output_fname)
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result
    return result


def form_fname(dir_name: str, fname_base: str, fname_ext: str) -> str:
    """
    form output filename from the given parameters
    Args:
      dir_name(str): Output directory or file name
      fname_base(str): If output directory given, then it should give output
        file's base name
      fname_ext(str): If output directory given, then it should give file's
        extension
    Returns:
      str: Output file name, if input is file name, then it is return as it is
        if the input is directory name, then output will be of the form
        <dir_name>/<fname_base>_<timestamp>.<fname_ext>
    """

    dir_path = Path(dir_name)

    ## If not directory then assume it is file name
    if not dir_path.is_dir():
        return dir_name

    dir_path = Path.absolute(dir_path)
    cur_time = int(datetime.timestamp(datetime.now()))
    output_fname = fname_base + "_" + str(cur_time) + "." + fname_ext
    output_fname = Path.joinpath(dir_path, output_fname)
    return output_fname


def args_check(args: Namespace) -> dict:
    """'
    check input arguments
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    ip_file = Path(args.input)
    if not ip_file.is_file():
        result["status"] = "Failure"
        result["msg"] = f"Input is not a valid file: {args.input}"
        return result

    op_path = Path(args.output)
    ## Output is either directory or file
    if (not op_path.is_dir()) and (not op_path.parent.is_dir()):
        result["status"] = "Failure"
        result["msg"] = f"Output path is not a valid path: {args.output}"
        return result

    if args.header < 1:
        result["status"] = "Failure"
        result["msg"] = f"Header Row cannot be less than 1: {args.header}"
        return result

    return result


def create_parser() -> Namespace:

    description = """ 
      Given input mastercomponent data file, extract unique manufacturer's
      name and create manufacturers records for them as per template file
      Note: For output option if *directory* name is given, then  generated 
      file will be saved as 'manufacturers_<timestamp>.xlsx in that directory. 
      If *filename" is given, then output will be saved in that name ( please
      provide proper path for the output )
    """

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "--input", "-i", type=str, required=True, help="Input master component file"
    )
    parser.add_argument(
        "--sheet",
        "-s",
        type=str,
        required=False,
        help="Input Sheet Name, by default first sheet data will be taken",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        required=False,
        default="../tmp",
        help="Path to destination directory or filename",
    )
    parser.add_argument(
        "--header",
        "-r",
        type=int,
        required=False,
        help="Input file Header Row Number",
        default=1,
    )
    return parser.parse_args()


if __name__ == "__main__":

    pp = pprint.PrettyPrinter(indent=2)

    ## Default output file name
    output_base = "COC"
    output_ext = "xlsx"

    args = create_parser()
    result = args_check(args)
    if result["status"] == "Success":
        output_fname = form_fname(args.output, output_base, output_ext)
        result = create_manufacturer(args, output_fname)

    print("Result Status: " + result["status"])
    if result["status"] == "Success":
        print("Output file is saved in " + str(output_fname))
    else:
        print("Error Message: " + str(result["msg"]))
