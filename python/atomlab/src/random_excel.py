from os import path
import random
import openpyxl as xl
from openpyxl.cell.cell import Cell
from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from pathlib import Path
from copy import copy
import argparse
from argparse import Namespace
import re


suffix_l = ["inc", "corp", "corporation", "llc", "ltd", "limited"]

# Function to convert string
def convert_string(raw_string):
    """
    Trim the string
    Convert all non-Alphanumeric characters to ""
    Convert to lowercase
    If the last word is any one of suffix_l then remove that ( useful
     to normalize manufacturer name )
    Reduce consequtive _ characters into single _
    """
    if not raw_string:
        return None
    tmp_s = str(raw_string).strip()
    tmp_s = re.sub("[^\w]+", " ", tmp_s)
    tmp_l = tmp_s.lower().split()
    if tmp_l[-1] in suffix_l:
        del tmp_l[-1]
    res = "".join(tmp_l)
    return res


# Function to copy styles of source cell into destination cell
def copy_cell_styles(dest_cell: Cell, source_cell: Cell) -> None:
    if source_cell.has_style:
        dest_cell.font = copy(source_cell.font)
        dest_cell.border = copy(source_cell.border)
        dest_cell.fill = copy(source_cell.fill)
        dest_cell.number_format = copy(source_cell.number_format)
        dest_cell.protection = copy(source_cell.protection)
        dest_cell.alignment = copy(source_cell.alignment)


# Freeze the first column (column at the left side of column B)
def freeze_first_column(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "B1"


# Freeze the first two columns (columns at the left side of column B)
def freeze_first_two_column(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "C1"


def get_coc_file_path(inp_file_path: Path, dest_file_path: Path) -> Path:
    # Get file name without extension
    coc_file_name = inp_file_path.stem

    # Get file extension
    coc_file_extension = inp_file_path.suffix

    # Create name for coc workbook file
    coc_file = f"{coc_file_name}_Coc{coc_file_extension}"

    # Generate coc workbook file path
    coc_file_path = dest_file_path.joinpath(coc_file)

    return coc_file_path


def get_manufacturer_file_path(inp_file_path: Path, dest_file_path: Path) -> Path:
    # Get file name without extension
    manufacturer_file_name = inp_file_path.stem

    # Get file extension
    manufacturer_file_extension = inp_file_path.suffix

    # Create name for manufacturer workbook file
    manufacturer_file = (
        f"{manufacturer_file_name}_Manufacturer{manufacturer_file_extension}"
    )

    # Generate manufacturer workbook file path
    manufacturer_file_path = dest_file_path.joinpath(manufacturer_file)

    return manufacturer_file_path


def get_master_file_path(inp_file_path: Path, dest_file_path: Path) -> Path:

    # Get file name without extension
    master_file_name = inp_file_path.stem

    # Get file extension
    master_file_extension = inp_file_path.suffix

    # Create name for master workbook file
    master_file = f"{master_file_name}_Master{master_file_extension}"

    # Generate master workbook file path
    master_file_path = dest_file_path.joinpath(master_file)

    return master_file_path


# Finds and return manufacturer column no in input worksheet
def get_manufacturer_column_no(inp_worksheet: Worksheet) -> int:
    manufacturer_column_no = None
    col_no = 1
    for cell in inp_worksheet[1]:
        if cell.value and cell.value.lower() == "manufacturer name":
            manufacturer_column_no = col_no
            break
        col_no += 1
    return manufacturer_column_no


# Finds and returns manufacturer part column no in input worksheet
def get_manufacturer_part_column_no(inp_worksheet: Worksheet) -> int:
    manufacturer_part_column_no = None
    col_no = 1
    for cell in inp_worksheet[1]:
        if cell.value and cell.value.lower() == "manufacturer part number":
            manufacturer_part_column_no = col_no
            break
        col_no += 1
    return manufacturer_part_column_no


def create_coc_workbook(inp_workbook: Workbook) -> Workbook:
    # Load the active worksheet of input workbook
    inp_worksheet = inp_workbook.active

    # Create a workbook for coc details
    coc_workbook = Workbook()

    # Load active worksheets of the coc workbook
    coc_worksheet = coc_workbook.active

    column_names = [
        "MFR Name",
        "MPN",
        "RoHS Status",
        "RoHS Version",
        "Exemption Annex",
        "RoHS Exemptions",
        "RoHS Source Doc Name",
        "RoHS Source Doc",
        "RoHS Doc Date",
        "RoHS Doc Type",
        "RoHS Doc Link",
        "REACH SVHC Status",
        "REACH SVHC Version",
        "REACH SVHC Source Doc Name",
        "REACH SVHC Source Doc",
        "REACH SVHC Doc Date",
        "REACH SVHC Doc Type",
        "REACH SVHC Doc Link",
        "REACH Annex XIV Version",
        "REACH Annex XIV Status",
        "REACH Annex XIV Source Doc Name",
        "REACH Annex XIV Source Doc",
        "REACH Annex XIV Doc Date",
        "REACH Annex XIV Doc Type",
        "REACH Annex XIV Doc Link",
        "REACH Annex XVII Version",
        "REACH Annex XVII Status",
        "REACH Annex XVII  Source Doc Name",
        "REACH Annex XVII Source Doc",
        "REACH Annex XVII Doc Date",
        "REACH Annex XVII Doc Type",
        "REACH Annex XVII Doc Link",
        "CA Pro 65 Version",
        "CA Pro 65 Status",
        "CA Pro 65 Potential Exposure",
        "CA Pro 65 Warning",
        "CA Pro 65 Source Doc Name",
        "CA Pro 65 Source Doc",
        "CA Pro 65 Doc Date",
        "CA Pro 65 Doc Type",
        "CA Pro 65 Doc Link",
        "TSCA Version",
        "TSCA Status",
        "TSCA Source Doc Name",
        "TSCA Source Doc",
        "TSCA Doc Date",
        "TSCA Doc Type",
        "TSCA Doc Link",
    ]

    # Add all the column names to the coc worksheet in the first row
    column_number = 1
    for column_name in column_names:
        current_cell = coc_worksheet.cell(row=1, column=column_number)
        current_cell.value = column_name
        current_cell.font = Font(bold=True)
        column_number += 1

    # Get column number of the manufacturer name in the input worksheet
    manufacturer_column_no = get_manufacturer_column_no(inp_worksheet)

    # Get column number of the manufacturer part no in the input worksheet
    manufacturer_part_column_no = get_manufacturer_part_column_no(inp_worksheet)

    # Create set of all manufacturer names and parth no from the input
    # worksheet
    manufacturer_names_part_numbers = set()
    row_no = 2
    while True:
        manufacturer_cell = inp_worksheet.cell(
            row=row_no, column=manufacturer_column_no
        )
        part_number_cell = inp_worksheet.cell(
            row=row_no, column=manufacturer_part_column_no
        )
        if not manufacturer_cell.value and not part_number_cell.value:
            break
        manufacturer = manufacturer_cell.value
        part_number = part_number_cell.value
        manufacturer_names_part_numbers.add((manufacturer, part_number))
        row_no += 1

    # Add all the acquired manufacturer names and parth no
    # to the manufacturer worksheet
    # in the 1st column
    dest_row_no = 2
    for manufacturer_name, part_number in manufacturer_names_part_numbers:
        dest_cell_1 = coc_worksheet.cell(row=dest_row_no, column=1)
        dest_cell_2 = coc_worksheet.cell(row=dest_row_no, column=2)
        dest_cell_1.value = manufacturer_name
        dest_cell_2.value = part_number
        dest_row_no += 1

    # Freeze the first two column of coc worksheet
    freeze_first_two_column(coc_worksheet)

    return coc_workbook


def create_manufacturer_workbook(inp_workbook: Workbook) -> Workbook:
    # Load the active worksheet of input workbook
    inp_worksheet = inp_workbook.active

    # Create a workbook for manufacturer details
    manufacturer_workbook = Workbook()

    # Load active worksheet of the manufacturer workbook
    manufacturer_worksheet = manufacturer_workbook.active

    column_names = [
        "Name",
        "category",
        "Aliases",
        "Type",
        "Contact Name",
        "Contact Email",
        "Contact Title",
        "contact_type",
        "address_line1",
        "address_line2",
        "city",
        "state",
        "postal_code",
        "country",
        "phone_number",
        "extension",
        "contact_extra_info",
        "Alert Mechanism",
        "Registration Name",
        "Registration Data",
        "Sender Email",
        "RSS Feed",
        "Alert Site",
    ]

    # Add all the column names to the manufacturer worksheet in the first row
    column_number = 1
    for column_name in column_names:
        current_cell = manufacturer_worksheet.cell(row=1, column=column_number)
        current_cell.value = column_name
        current_cell.font = Font(bold=True)
        column_number += 1

    # Get column number of the manufacturer name in the input worksheet
    manufacturer_column_no = get_manufacturer_column_no(inp_worksheet)

    # Create set of all manufacturer names from the input worksheet
    manufacturer_names = set()
    row_no = 2
    while True:
        source_cell = inp_worksheet.cell(row=row_no, column=manufacturer_column_no)
        if not source_cell.value:
            break
        manufacturer = source_cell.value
        manufacturer_names.add(manufacturer)
        row_no += 1

    # Add all the acquired manufacturer names to the manufacturer worksheet
    # in the 1st column
    dest_row_no = 2
    for manufacturer_name in manufacturer_names:
        dest_cell = manufacturer_worksheet.cell(row=dest_row_no, column=1)
        dest_cell.value = manufacturer_name
        dest_row_no += 1

    # Freeze first column of the manufacturer worksheet
    freeze_first_column(manufacturer_worksheet)

    return manufacturer_workbook


def create_master_workbook(inp_worksheet: Worksheet, args: Namespace) -> Workbook:

    # ----------- Load start & end values of row & columns -------------------

    # Load start & end row values
    start_row = args.sr
    end_row = args.er

    # Load start & end column values
    start_col = args.sc
    end_col = args.ec

    # Defaults to 2 if start row value is empty
    # [Since 1st row is used for column names]
    start_row = int(start_row) if start_row.strip() else 2
    # Defaults to 1 if start column value is empty
    start_col = int(start_col) if start_col.strip() else 1

    # Defaults to max value if any end row or column values are empty
    end_row = int(end_row) if end_row.strip() else inp_worksheet.max_row
    end_col = int(end_col) if end_col.strip() else inp_worksheet.max_column
    # ------------------------------------------------------------------------

    # Create a workbook to store desired output
    master_workbook = Workbook()

    # Load active worksheet of the master workbook
    master_worksheet = master_workbook.active

    # Iterate through 1st row till given end column in the source worksheet
    # and get the column cells
    first_row = inp_worksheet[1][:end_col]
    column_headers = [cell for cell in first_row if cell.value]

    # Insert column values and styles into the 1st row
    # of the master workbook
    column_number = 1
    for column_cell in column_headers:
        current_cell = master_worksheet.cell(row=1, column=column_number)
        current_cell.value = column_cell.value
        copy_cell_styles(current_cell, column_cell)
        column_number += 1

    # Get column number of the manufacturer name in the input worksheet
    manufacturer_column_no = get_manufacturer_column_no(inp_worksheet)

    # Generate random numbers within given range of row values
    # To include the last row (end row) -> + 1 is added
    total_rows = end_row - start_row + 1
    population = range(start_row, end_row + 1)
    randomized_row_values = random.sample(population, total_rows)

    # Copying randomized source rows within given range into dest worksheet
    dest_row_no = 2  # Since 1st row already contains column names
    for source_row_no in randomized_row_values:
        for column_no in range(start_col, end_col + 1):
            # Get current source and destination cell
            current_dest_cell = master_worksheet.cell(row=dest_row_no, column=column_no)
            current_source_cell = inp_worksheet.cell(
                row=source_row_no, column=column_no
            )
            # Copy value and style of source cell into destination cell
            if column_no == manufacturer_column_no:
                current_dest_cell.value = convert_string(current_source_cell.value)
            else:
                current_dest_cell.value = current_source_cell.value
            copy_cell_styles(current_dest_cell, current_source_cell)
        dest_row_no += 1

    return master_workbook


def main(args: Namespace) -> None:
    # Load source excel file path
    source_file_path = Path(args.inp)

    # Load destination excel file path
    dest_file_path = Path(args.out)

    # Load the source excel workbook
    source_workbook = xl.load_workbook(source_file_path)

    # Load active worksheet of the workbook
    source_worksheet = source_workbook.active

    # Create and save master workbook
    master_workbook = create_master_workbook(source_worksheet, args)
    master_file_path = get_master_file_path(source_file_path, dest_file_path)
    master_workbook.save(master_file_path)

    # Create and save coc workbook
    coc_workbook = create_coc_workbook(master_workbook)
    coc_file_path = get_coc_file_path(source_file_path, dest_file_path)
    coc_workbook.save(coc_file_path)

    # Create and save manufacturer workbook
    manufacturer_workbook = create_manufacturer_workbook(master_workbook)
    manufacturer_file_path = get_manufacturer_file_path(
        source_file_path, dest_file_path
    )
    manufacturer_workbook.save(manufacturer_file_path)


def create_parser() -> Namespace:

    example = """
    Example:
    python random_excel.py --inp="Path to input excel file" --out="Path to destination folder" --sr=10 --er=20 --sc=1 --ec=
    NOTE: Start row --sr cannot be given as 1 
    because it will take header row as input.
    we need to give 2 as input if we need 1st row.
    """
    parser = argparse.ArgumentParser(description=example)

    parser.add_argument("--inp", help="Input excel file path")
    parser.add_argument("--out", help="Path to destination folder to save output")
    parser.add_argument("--sr", help="Start row number")
    parser.add_argument("--er", help="End row number")
    parser.add_argument("--sc", help="Start column number")
    parser.add_argument("--ec", help="End column number")

    return parser.parse_args()


if __name__ == "__main__":
    args = create_parser()
    main(args)
