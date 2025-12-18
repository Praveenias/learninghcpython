###############################################################################
# program to split single xlsx files into multiple xlsx files using openpyxl.
###############################################################################

import math
import pprint
import argparse
import pathlib
from openpyxl import load_workbook
from openpyxl import Workbook
status={}

def check_input(ip_fname,no_of_files=0,rows_per_file=0):
    result = {
    'status':'Success',
    'msg'   : '',
    }

    ip_path = pathlib.Path(ip_fname) 
    if ( not ip_path.is_file()):
        result['status'] = 'Failure'
        result['msg'] = (f'Input is not a valid file: {ip_fname}')
        return result

    ## Eiter one of the rows_per_file , no_of_files should be given
    if ( (rows_per_file < 1) and (no_of_files < 1) ):
        result['status'] = 'Failure'
        tmp ='Both rows_per_file and no_of_files cannot be less than 1: ' 
        result['msg'] = (f'{tmp} {rows_per_file}, {no_of_files}')
        return result

def extractdata(start,end,ip_sheetname,i,header_l):
    try:
        op_workbook=Workbook()
        op_worksheet=op_workbook.active
        op_worksheet.append(header_l)
        for row in ip_sheetname.iter_rows(min_row=start+1,max_row=end):
            list1=[cell.value for cell in row]
            op_worksheet.append(list1)
        op_workbook.save(pathlib.Path(ip_fname).stem+"_"+str(i)+".xlsx")
    except Exception as err:
        result['status'] = 'Failure'
        result['msg'] = err
        return result

def splitdata(ip_fname,ip_sheetname,no_of_files=0,rows_per_file=0):
    result = {
      'status':'Success',
      'msg':'',
    }
    try:
        ip_workbook=load_workbook(ip_fname)
        if (ip_sheetname):
            ip_sheetname = ip_workbook[ip_sheetname]
        else:
            ip_sheetname = ip_workbook.active
        no_iprows= ip_sheetname.max_row-1
        header_l =[cell.value for cell in ip_sheetname[1]]
        start = 1
        if no_of_files:
            end = math.ceil(no_iprows/no_of_files)
        if rows_per_file:
            no_of_files = math.ceil(no_iprows/rows_per_file)
            end = rows_per_file
        for i in range(1,no_of_files+1):
            extractdata(start,end+start,ip_sheetname,i, header_l)
            start+=end
    except Exception as err:
        result['status'] = 'Failure'
        result['msg'] = err
        return result
    return result


#############################################
#   Start of main Program
#############################################
pp = pprint.PrettyPrinter(indent=2)

## Command Line Processing
cli_parser = argparse.ArgumentParser(
    prog='split_excel',
    description='Split given work sheet into multiple files')

# Add the arguments
cli_parser.add_argument('--input', '-i', type=str, required=True,
                        help='Input Excel file')

cli_parser.add_argument('--sheet', '-s', type=str, required=False,
                help="Input and Output Sheet Name")

cli_parser.add_argument('--nofiles', '-f', type=int, required=False,
                        help='Number of Files')

cli_parser.add_argument('--norows', '-r', type=int, required=False,
                        help='rows per file')

args = cli_parser.parse_args()

ip_fname       = args.input
ip_sheetname   = args.sheet
no_of_files    = args.nofiles
rows_per_file  = args.norows

result=check_input(ip_fname,no_of_files=no_of_files,rows_per_file=rows_per_file)
print('Start of Excel Split')
result=splitdata(ip_fname,ip_sheetname,no_of_files=no_of_files,rows_per_file=rows_per_file)
pp.pprint(result)
