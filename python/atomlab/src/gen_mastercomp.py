"""
  Given any excel file contain large number of records, it extracts given
  number of random row from that and store it is output file.
  Useful to get subset of Master Components from large set of master component
  data
"""

import pprint
from datetime import datetime
from pathlib import Path
import argparse
from argparse import Namespace
import random

import openpyxl as xl

## To write Excel work book
from openpyxl.workbook import workbook


def create_master(args: Namespace, output_fname: str) -> dict:
    """
    Create random records of Mastercomponent excel file from given
    input file
    Args:
      args(dictionary):
        input : Input master component file name ( with path )
        sheet : Sheet name , containing input data ( if not given first sheet )
        header : Header row number. Data assumed to one row after header till
                  end of the file
        number : Number of records in output file
    Returns:
     Dictionay
      status: 'Success' or 'Failure'
      msg:     If failure will contain error message else empty string
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    try:
        # ~ If we set read only then hyper link is not read
        ip_workbook = xl.load_workbook(filename=args.input, data_only=True)
        if args.sheet:
            ip_sheet = ip_workbook[args.sheet]
        else:
            ip_sheet = ip_workbook.active
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result

    population = range((args.header + 1), ip_sheet.max_row)
    if args.number >= len(population):
        result["status"] = "Failure"
        tmp = "Output Number of rows is greater than number of rows in "
        tmp = tmp + " input file. "
        tmp1 = "Input no.of rows: {ip_sheet.max_rows}"
        tmp2 = "Output no.of rows: {args.number}"
        result["msg"] = f"{tmp} {tmp1} {tmp2}"
        return result

    row_numbers = random.sample(population, args.number)

    master_workbook = xl.Workbook()
    master_worksheet = master_workbook.active
    master_worksheet.title = "Data"

    try:
        ## First append header row
        master_worksheet.append(cell.value for cell in ip_sheet[args.header])
        for i in row_numbers:
            master_worksheet.append(cell.value for cell in ip_sheet[i])
        master_workbook.save(output_fname)
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result

    return result


def form_fname(dir_name: str, fname_base: str, fname_ext: str) -> str:
    """
    form output filename from the given parameters
    Args:
      dir_name(str): Output directory or file name
      fname_base(str): If output directory given, then it should give output
        file's base name
      fname_ext(str): If output directory given, then it should give file's
        extension
    Returns:
      str: Output file name, if input is file name, then it is return as it is
        if the input is directory name, then output will be of the form
        <dir_name>/<fname_base>_<timestamp>.<fname_ext>
    """

    dir_path = Path(dir_name)

    ## If not directory then assume it is file name
    if not dir_path.is_dir():
        return dir_name

    dir_path = Path.absolute(dir_path)
    cur_time = int(datetime.timestamp(datetime.now()))
    output_fname = fname_base + "_" + str(cur_time) + "." + fname_ext
    output_fname = Path.joinpath(dir_path, output_fname)
    return output_fname


def args_check(args: Namespace) -> dict:
    """'
    check input arguments
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    ip_file = Path(args.input)
    if not ip_file.is_file():
        result["status"] = "Failure"
        result["msg"] = f"Input is not a valid file: {args.input}"
        return result

    op_path = Path(args.output)
    ## Output is either directory or file
    if (not op_path.is_dir()) and (not op_path.parent.is_dir()):
        result["status"] = "Failure"
        result["msg"] = f"Output path is not a valid path: {args.output}"
        return result

    if args.header < 1:
        result["status"] = "Failure"
        result["msg"] = f"Header Row cannot be less than 1: {args.header}"
        return result

    if args.number < 1:
        result["status"] = "Failure"
        tmp = "Number of Rows to extract cannot be less than 1: "
        result["msg"] = f"{tmp}: {args.number}"
        return result
    return result


def create_parser() -> Namespace:

    description = """ 
      Given input excel data file, extract given number of random rows 
      and store the same in output file. Note: For output option if *directory* 
      name is given, then  generated file will be saved as 
      'mastercomp_random_<timestamp>.xlsx in that directory.  If *filename" 
      is given, then output will be saved in that name ( please provide proper
      path for the output )
    """

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "--input", "-i", type=str, required=True, help="Input excel file"
    )
    parser.add_argument(
        "--sheet",
        "-s",
        type=str,
        required=False,
        help="Input Sheet Name, by default first sheet data will be taken",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        required=False,
        default="../tmp",
        help="Path to destination directory or filename",
    )
    parser.add_argument(
        "--number",
        "-n",
        type=int,
        required=False,
        help="Number of Records to extract",
        default=1,
    )
    parser.add_argument(
        "--header",
        "-r",
        type=int,
        required=False,
        help="Header Row Number. Note: Excel row number starts with 1",
        default=1,
    )
    return parser.parse_args()


if __name__ == "__main__":

    pp = pprint.PrettyPrinter(indent=2)

    ## Default output file name
    output_base = "mastercomp_random"
    output_ext = "xlsx"

    args = create_parser()
    result = args_check(args)
    if result["status"] == "Success":
        output_fname = form_fname(args.output, output_base, output_ext)
        result = create_master(args, output_fname)

    print("Result Status: " + result["status"])
    if result["status"] == "Success":
        print("Output file is saved in " + str(output_fname))
    else:
        print("Error Message: " + result["msg"])
