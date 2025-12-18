from pathlib import Path
import json
import openpyxl
import re
from mysql.connector import connect, Error

        
class UploadFile:

  def __init__(self,file_type="manufacturer"):
    config_file_name='config.json'
    self.config_data = self.get_config_data(config_file_name)
    if not self.config_data:
      return None
    self.db_schema = self.config_data["db_schema"]
    self.db_config = self.config_data["db_config"]
    #print(self.db_schema)
    

  def get_config_data(self,file_name:str):
    try:
      with open(Path(file_name),encoding='utf-8') as data:
        return json.load(data)
    except Exception as e:
      print(e)
      return {}
  def main():
    pass

  def normalize_string(self,raw_string:str,remove_suffix:bool=False)->str:
    suffixL = ['inc', 'corp', 'corporation', 'llc', 'ltd', 'limited']
    if (not raw_string):
      return None
    tmp_s = str(raw_string).strip()
    tmp_s = re.sub('[^\w]+', ' ',tmp_s)
    tmp_l = tmp_s.lower().split()
    if (tmp_l[-1] in suffixL) and remove_suffix: del tmp_l[-1]
    res = '_'.join(tmp_l)
    return res

  def get_header_data(self,header_row,ws):
    header_names = [cell.value for cell in ws[header_row]]
    if None in header_names:
      none_index = header_names.index(None)
      header_names = header_names[:none_index]
    normalized_header = list(map(self.normalize_string, header_names))
    return normalized_header

  def get_column_index(self,header:list,column) ->dict:
    column_index = {}
    for col in column:
      column_index[col]=header.index(col)
    return column_index
  
  def add_remining_column_values(self,extra_column,values):
    for column in extra_column:
      if column == "normalized_name":values["normalized_name"] = self.normalize_string(values["name"],True)

    return values


  def upload_manufacturer_data(self,file_data):

    #basic manufacturer table information
    manufacturer_db_schema = self.db_schema[file_data["file_type"]]
    #table_name = manufacturer_db_schema['table_name']
    column_names = tuple(manufacturer_db_schema['columns'].keys())
    sql_insert_query = manufacturer_db_schema["insert_query"]

    #loading excel
    ipWorkBook = openpyxl.load_workbook(filename= file_data['file_path'])
    ipworksheet = ipWorkBook.active

    #getting header value
    header_column = self.get_header_data(file_data['header_row'],ipworksheet)

    #checking column names are missing in excel
    if not all(item in header_column for item in file_data['column_names']):
      header_missing = set(file_data['column_names'])-set(header_column)
      print(f'{header_missing}: Mandatory Column name(s) are missing')
      return False

    #getting mandatory columns index value in header list
    column_index = self.get_column_index(header_column,file_data['column_names'])

    try:
      with connect(**self.db_config) as connection:
        with connection.cursor() as cursor:
          for val in self.get_manufacturer_data_from_excel(ipworksheet,file_data,column_index,manufacturer_db_schema):
            if val == None:
              break
            val = self.add_remining_column_values(file_data['columns_not_in_template'],val)
            cursor.execute(sql_insert_query,tuple(val.values()))
            print(val)
        connection.commit()
    except Error as e:
      print(e)

  #generator_function
  def get_manufacturer_data_from_excel(self,ipworksheet,file_data,column_index,manufacturer_db_schema):
    #basic manufacturer excel information
    min_rows = file_data['header_row']+1
    max_rows = ipworksheet.max_row

    for data in ipworksheet.iter_rows(min_row=min_rows,max_row=max_rows,values_only=True ):
      if set(data) == {None}:
        yield None
      for col in file_data['column_names']:
        manufacturer_db_schema['columns'][col] = data[column_index[col]]
      yield manufacturer_db_schema['columns']


  def load_file_date(self):
    for file in self.config_data['file']:

      #manufacturer file , manufacturer table
      if file["file_type"] == 'manufacturer':
        self.upload_manufacturer_data(file)

      #master file ,  master+compliance_results table
      if file["file_type"] == 'master':
        self.upload_manufacturer_data(file) 
        pass


if __name__ == '__main__':
  uf = UploadFile()
  out = uf.load_file_date()
  #print(config_data['db'])


  # {
	# 		"file_type":"manufacturer",
	# 		"file_path":"manufacturer.xlsx",
	# 		"header_row":2,
	# 		"column_names":["name","category","alert_mechanism"],
	# 		"columns_not_in_template":["normalized_name"]
	# 	},
