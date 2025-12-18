#permutation and combination

from openpyxl import Workbook

combinations = {
  "part_life_cycle_status":["active","eol","nrnd","unavailable","unconfirmed"],
  "rohs_compliance" :["Compliant","Non-compliant","c_w_e_pastdate","c_w_e_futuredate","NA"],
  "reach_compliance" : ["Compliant","Non-compliant","NA"], #Compliant-with-exemption
  "man_stock_risk" : ["man_high","man_low","man_empty"],
  "sup_stock_risk" : ["sup_low","sup_high","sup_empty"],
  "man_leadtime_risk" : ["man_high","man_low","man_empty"],
  "sup_leadtime_risk" : ["sup_low","sup_high","sup_empty"]
}

man_stock_map = {"man_high":1500,"man_low":100,"man_empty":''}
sup_stock_map = {"sup_high":1500,"sup_low":100,"sup_empty":''}
man_leadtime_map = {"man_high":50,"man_low":12,"man_empty":''}
sup_leadtime_map = {"sup_high":50,"sup_low":12,"sup_empty":''}

man_template = {'Name':'Intel', 'Author Emails':'', 'Approver Emails':'', 'category':'public', 'Aliases':'', 'Type':'', 'Contact Name':'', 'Contact Email':'', 'Contact Type':'', 'Address Line1':'', 'Alert Mechanism':'Manual', 'Registration Name':'oemalerts@ilensys.com', 'Registration Data':'', 'Sender Email':'', 'RSS Feed':'', 'Alert Site':'', 'reference':'', 'Global Name':'', 'Code':'', 'Rank':'', 'Web Site':'', 'Notes':''}

mas_template = {
  'Manufacturer Name':'', 'Manufacturer Part Number':'', 'Generic Part Number':'', 'Manufacturer Part Description':'desc1', 'Part Category':'Electrical & Electronics', 'Component Type':'Capacitor', 'Part Type':'Array Diode', 'Assembly Type':' SMT', 'Power Type':'Active', 'Regulatory Test':'', 'Part Lifecycle Status':'', 'Stock Details':'', 'Lead Time':'', 'Data Sheet Document':'', 'Price($)':'', 'Component Classification':'', 'Part_Mass':'', 'UOM':'', 'Year to EOL':'', 'RoHS3':'', 'RoHS3 Comment':'', 'RoHS3 Exemption Date':'', 'REACH':'', 'Check Date Regulatory':'', 'check Date Lifecycle':'', 'Check Date Stock':'', 'Check Date Compliance':'', 'LTB Possibility':'', 'Remark':''
}

bom_template = {
  'Customer Part number':'', 'Part Revision':'', 'Level':'', 'Customer part number description':'', 'Manufacturer name':'', 'Manufacturer part number':'', 'Quantity':'1', 'Proactive Analysis':'Yes', 'Total Inventory':'', 'EAU':'', 'Critical Component':'', 'Supplier Internal Partnumber':'', 'Supplier Partnumber':'', 'Supplier Manufacturer':'', 'Supplier Lead Time':'', 'Supplier Stock':'', 'Supplier Monthly Usage':'', 'Supplier Cover Months':'', 'Supplier Cover Date':'', 'Check Date Supplier':'', 'Check Date life cycle':'', 'Check Date Stock':'', 'LTB Possibility':'', 'OPO':'', 'Remark':''
}

supplier_template = {
    'Company Name':'Dell','Division Name':'Dell div1','BU Name':'Dell site1 ',
    'Product Name':'Dell Laptop','Product Model Name':'M1',
    'BOM Name':'sanity_check2','Customer Part number':'','ATOM - Manufacturer name':'Intel',	
    'ATOM - Manufacturer partnumber':''	,'Supplier Lead Time':'',
    'Supplier Cover Date':'','Check Date Supplier':'','OPO':''
}






combination_count = 1
for i in combinations.keys():
  combination_count *= len(combinations[i])
print("combination_count : ",combination_count)

#manufacturer_file
man_workbook = Workbook()
man_sheet = man_workbook.active
man_sheet.title = "Data"

#master_file
mas_workbook = Workbook()
mas_sheet = mas_workbook.active
mas_sheet.title = "Data"

#bom_file
bom_workbook = Workbook()
bom_sheet = bom_workbook.active
bom_sheet.title = "Data"

#Supplier_file
supplier_workbook = Workbook()
supplier_sheet = supplier_workbook.active
supplier_sheet.title = "Data"

#generator_function_to_get_combination
def get_combination():
  for plcs in combinations["part_life_cycle_status"]:
    for rec in combinations["reach_compliance"]:
      for roc in combinations["rohs_compliance"]:
        for msr in combinations["man_stock_risk"]:
          for ssr in combinations["sup_stock_risk"]:
            for mltr in combinations["man_leadtime_risk"]:
              for sltr in combinations["sup_leadtime_risk"]:
                comb = {"plcs":plcs,"rec":rec,"roc":roc,"msr":msr,"ssr":ssr,"mltr":mltr,"sltr":sltr}
                yield comb

#manufacturer file
man_sheet.append([data for data in man_template.keys()])
man_sheet.append([data for data in man_template.values()])
man_workbook.save('manufacturer.xlsx')

def get_rohs_data(rohs_d)->dict:
  if rohs_d in ["Compliant","Non-compliant","NA"]:
    return {"rohs":rohs_d,"date":''}
  rohs_dict = {"rohs":"Compliant-with-exemption","date":"2021-02-02"}
  if rohs_d == "c_w_e_futuredate":
    rohs_dict["date"]="2025-02-02"
  return rohs_dict

  
    

def load_master(mpn,data):
  mas_template['Manufacturer Name'] = "Intel"
  mas_template['Manufacturer Part Number'] = mpn
  mas_template['Manufacturer Part Description'] = ",".join([i for i in data.values()])
  mas_template["Part Lifecycle Status"] = data["plcs"]
  mas_template["Stock Details"] =  man_stock_map[data['msr']]
  mas_template["Lead Time"] = man_leadtime_map[data['mltr']]
  rohs_dict = get_rohs_data(data["roc"])
  mas_template["RoHS3"] = rohs_dict["rohs"]
  mas_template["RoHS3 Exemption Date"] = rohs_dict["date"]
  mas_template["REACH"] = data["rec"]

  mas_data = [data for data in mas_template.values()]
  mas_sheet.append(mas_data)

def load_bom(mpn,data):
  bom_template["Customer Part number"] = mpn
  bom_template["Customer part number description"] = ",".join([i for i in data.values()])
  bom_template["Manufacturer name"] = "Intel"
  bom_template["Manufacturer part number"] = mpn
  bom_template["Supplier Stock"] = sup_stock_map[data['ssr']]
  bom_template["Supplier Lead Time"] = sup_leadtime_map[data['sltr']]
  bom_data = [data for data in bom_template.values()]
  bom_sheet.append(bom_data)
  
def load_supplier(mpn):
  supplier_template['ATOM - Manufacturer partnumber']=mpn
  supplier_template['Supplier Lead Time']=12
  supplier_template['Customer Part number']=mpn
  supplier_template['Supplier Cover Date']="12-07-2024"
  supplier_template['Check Date Supplier']="12-07-2023"
  supplier_template['OPO']=100
  supplier_data = [data for data in supplier_template.values()]
  supplier_sheet.append(supplier_data)


mas_sheet.append([data for data in mas_template.keys()])
bom_sheet.append([data for data in bom_template.keys()])
supplier_sheet.append([data for data in supplier_template.keys()])

mpn=10000
for data in get_combination():
  manufacturer_part_number = mpn
  #load_master(manufacturer_part_number,data)
  #load_bom(manufacturer_part_number,data)
  load_supplier(manufacturer_part_number)
  mpn +=1
  if mpn == 14999:break

#mas_workbook.save("master.xlsx")
#bom_workbook.save("bom.xlsx")
supplier_workbook.save("supplier.xlsx")

#print(get_combination())

#workbook.save("12.xlsx")

    
  
  
