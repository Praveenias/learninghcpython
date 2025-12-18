from openpyxl import load_workbook

wb = load_workbook('bom_s.xlsx')
ws = wb.active
print([cell.value for cell in ws[2]])

