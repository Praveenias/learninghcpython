"""
  Given any excel file contain mastercomponents,
  it creates SCIP Article data file as required
  by the ATOM template
  # python Scip_Articledata.py -i ../tmp/25/mastercomp_random_1646725978.xlsx -ir ../tmp/25/ref.xlsx --number 25  
"""
import pprint
from datetime import datetime
from pathlib import Path
import argparse
from argparse import Namespace
import re
from faker import Faker
import random
import traceback

import openpyxl as xl

## To write Excel work book
from openpyxl.workbook import workbook

column_names = [
    "Manufacturer Name",
    "Manufacturer part number",
    "Level",
    "Number of Units (quantity)",
    "Is Complex Object",
    "Article Name",
    "Other article Names",
    "Primary article identifier type",
    "Primary article identifier value",
    "Other article identifiers",
    "Article Category Identifier",
    "Production in European Union",
    "Height",
    "Height_uom",
    "Length",
    "Length_uom",
    "Width",
    "Width_uom",
    "Diameter",
    "Diameter_uom",
    "Density",
    "Density_uom",
    "Weight",
    "Weight_uom",
    "Volume",
    "Volume_uom",
    "Colour",
    "Other Characteristics",
    "Safe use instruction(s)",
    "Substance Name",
    "CAS",
    "Concentration Range",
    "Material Category",
    "Additional material characteristics",
    "Mixture category(EUPCS)",
    "Candidate list substance no longer present",
]


suffix_l = ["inc", "corp", "corporation", "llc", "ltd", "limited"]


def normalize_string(raw_string):
    """
    Trim the string
    Convert all non-Alphanumeric characters to _
    Convert to lowercase
    If the last word is any one of suffix_l then remove that ( useful
     to normalize manufacturer name )
    Reduce consequtive _ characters into single _
    """
    if not raw_string:
        return None
    tmp_s = str(raw_string).strip()
    tmp_s = re.sub("[^\w]+", " ", tmp_s)
    tmp_l = tmp_s.lower().split()
    if tmp_l[-1] in suffix_l:
        del tmp_l[-1]
    res = "_".join(tmp_l)
    return res


k = -1


def create_levels(max_rows: int = 20) -> list[int]:
    base_level = 1

    max_sub_levels_allowed = 5

    row = [base_level]
    res = [(str(base_level), "yes")]
    for i in range(max_rows):

        isComplexObject = (
            "Yes" if random.randint(1, 10) / 10 + 0.85 > 1 else "No"
        )

        if len(row) <= 1 and isComplexObject == "No":
            row = [row[0]]
            continue

        if (
            i == max_rows - 1
            or len(row) == max_sub_levels_allowed
            or (len(row) > 1 and row[-1] > 5)
        ):
            isComplexObject = "No"

        if isComplexObject == "Yes":
            row.append(1)
        else:
            y = row.pop()
            try:
                row[-1] += 1
            except:
                row = [y + 1]
        display_message = f'{".".join(map(str, row))}'  # {isComplexObject}
        # print(display_message)
        res.append((display_message, isComplexObject))
    return res


# def create_levels(max_rows: int = 20) -> list[int]:
#     base_level = 2

#     max_sub_levels_allowed = 4
#     row = [base_level]
#     res = [(str(base_level), "yes")]
#     for i in range(max_rows):
#         isComplexObject = "Yes" if random.randint(1, 10) / 10 + 0.85 > 1 else "No"

#         if len(row) <= 1 and isComplexObject == "No":
#             row = [row[0] + 1]
#             continue

#         if (
#             i == max_rows - 1
#             or len(row) == max_sub_levels_allowed
#             or (len(row) > 1 and row[-1] > 4)
#         ):
#             isComplexObject = "No"

#         if isComplexObject == "Yes":
#             row.append(1)
#         else:
#             y = row.pop()
#             try:
#                 row[-1] += 1
#             except:
#                 row = [y + 1]
#         display_message = f'{".".join(map(str, row))}'  # {isComplexObject}
#         # print(display_message)
#         res.append((display_message, isComplexObject))
#     return res


def find_Substance_column(ip_sheet1: xl.worksheet, hdr_row_no: int) -> int:

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet1[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("substancename")
    except ValueError:
        column_number = -1
    return column_number


def find_CAS_Number_column(ip_sheet1: xl.worksheet, hdr_row_no: int) -> int:

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet1[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("casnumber")
    except ValueError:
        column_number = -1
    return column_number


def find_mfr_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("manufacturername")
    except ValueError:
        column_number = -1
    return column_number


def find_mpn_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("manufacturerpartnumber")
    except ValueError:
        column_number = -1
    return column_number


def find_ibom_mfr_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        if cell.value:
            val = cell.value.strip().lower()
            val = re.sub(pat, "", val)
        else:
            val = None
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("mappedmanufacturername")
    except ValueError:
        column_number = -1
    return column_number


def find_ibom_mpn_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        if cell.value:
            val = cell.value.strip().lower()
            val = re.sub(pat, "", val)
        else:
            val = None
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("importedmanufacturerpartnumber")
    except ValueError:
        column_number = -1
    return column_number


def get_Substance_Name(
    ip_sheet1: xl.worksheet,
    hdr_row_no: int,
    Substance_column: int,
    max_row_no: int,
) -> list:
    st_row = hdr_row_no + 1
    Substance_Name = [
        ip_sheet1[r][Substance_column].value
        for r in range(st_row, max_row_no + 1)
    ]
    return Substance_Name


def get_CAS_Number(
    ip_sheet1: xl.worksheet,
    hdr_row_no: int,
    CAS_Number_column: int,
    max_row_no: int,
) -> list:
    st_row = hdr_row_no + 1
    CAS_Number = [
        ip_sheet1[r][CAS_Number_column].value
        for r in range(st_row, max_row_no + 1)
    ]
    return CAS_Number


def get_mfr_names(
    ip_sheet: xl.worksheet, hdr_row_no: int, mfr_col: int, max_row_no: int
) -> list:
    st_row = hdr_row_no + 1
    mfr_names = [
        ip_sheet[r][mfr_col].value for r in range(st_row, max_row_no + 1)
    ]
    return mfr_names


def get_mpn_names(
    ip_sheet: xl.worksheet, hdr_row_no: int, mpn_col: int, max_row_no: int
) -> list:
    st_row = hdr_row_no + 1
    mpn_names = [
        ip_sheet[r][mpn_col].value for r in range(st_row, max_row_no + 1)
    ]
    return mpn_names


def get_mfr_record(
    mfr: str, mpn: str, Substance: str, CAS: str, max_rows: int
) -> list:
    """
    ToDo: Need to fill up all the mandatory column
    """
    global k
    k += 1
    Number_of_Units_quantity = "1"
    Candidate_list_substance_no_longer_present = ("No",)
    Concentration_Range = ("≥ 0.3% w/w and < 1.0% w/w",)
    Other_article_Names = ("Brand:HP, Model:Elitebook",)
    Production_in_European_Union = ("EU produced",)
    Safe_use_instructions = (
        "Lorem ipsum dolor sit amet, consectetur adipiscing elit."
    )
    Primary_article_identifier_type = [
        random.choice(["part number", "catalogue number"])
    ]
    # Level = [
    #     "1",
    #     "1.1",
    #     "1.1.1",
    #     "1.1.1.1",
    #     "1.1.1.1.1",
    #     "1.1.1.2",
    #     "1.1.2",
    #     "1.1.2.1",
    #     "1.1.2.1.1",
    #     "1.2",
    #     "1.2.1",
    #     "1.2.1.1",
    #     "1.2.1.1.1",
    #     "1.2.1.2",
    #     "1.2.1.2.1",
    #     "1.3",
    #     "1.3.1",
    #     "1.4",
    #     "1.4.1",
    # ]
    # Level = Level[k % 19]
    # Level = (
    #     str(int(Level.split(".")[0]) + k // 19) + "." + ".".join(Level.split(".")[1:])
    # )

    # is_complex_object = [
    #     "Yes",
    #     "Yes",
    #     "Yes",
    #     "Yes",
    #     "No",
    #     "No",
    #     "Yes",
    #     "Yes",
    #     "No",
    #     "Yes",
    #     "Yes",
    #     "Yes",
    #     "No",
    #     "Yes",
    #     "No",
    #     "Yes",
    #     "No",
    #     "Yes",
    #     "No",
    # ]
    # is_complex_object = is_complex_object[k % 19]
    Artilce_Name = [
        "Laptop",
        "Monitor",
        "Board-1",
        "Resitor",
        "Die",
        "Resitor",
        "Board-2",
        "Resitor",
        "Die",
        "Keyboard",
        "Board3",
        "Resitor",
        "Die",
        "Blade",
        "Metal",
        "Motherboard",
        "Die",
        "Keyboard",
        "Metal",
    ]
    Artilce_Name = Artilce_Name[k % 19]
    Material_Category = [
        "66367",
        "66368",
        "66369",
        "66370",
        "66371",
        "66372",
        "66373",
        "66374",
        "66375",
        "66376",
        "66377",
        "66378",
        "66379",
        "66376",
    ]
    Material_Category = Material_Category[k % 14]
    Additional_material_characteristics = [
        "66626",
        "66642",
        "66633",
        "66651",
        "66621",
        "66644",
        "66645",
        "66624",
        "66630",
        "66638",
        "66627",
        "66632",
        "66626",
        "66638",
    ]
    Additional_material_characteristics = Additional_material_characteristics[
        k % 14
    ]
    Other_Characteristics = [
        "Operating Temperature:-40°C ~ 220°C, Resistance:1.5 kOhms",
        None,
        None,
        "Operating Temperature:-40°C ~ 220°C, Resistance:1.5 kOhms",
        None,
        None,
        None,
        "Operating Temperature:-40°C ~ 220°C, Resistance:1.5 kOhms",
        None,
        None,
        None,
        None,
        None,
        None,
    ]
    Other_Characteristics = Other_Characteristics[k % 14]
    Colour = [
        "64912,64913",
        "64912,64914",
        "64912,64915",
        "64912,64916",
        "64912,64917",
        "64912,64918",
        "64912,64919",
        "64912,64920",
        "64912,64921",
        "64912,64922",
        "64912,64923",
        "64912,64924",
        "64912,64925",
        "64912,64922",
    ]
    Colour = Colour[k % 14]
    Volume_uom = ("dm³",)
    Volume = "1"
    Weight_uom = "g"
    Weight = [
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "2",
        "3",
        "4",
        "5",
        "2",
    ]
    Weight = Weight[k % 14]
    Density_uom = ("kg/m³",)
    Density = [
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "2",
        "3",
        "4",
        "5",
        "2",
    ]
    Density = Density[k % 14]
    Width_uom = "m"
    Width = [
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "2",
        "3",
        "4",
        "5",
        "2",
    ]
    Width = Width[k % 14]
    Diameter = [
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "2",
        "3",
        "4",
        "5",
        "2",
    ]
    Diameter = Diameter[k % 14]
    Diameter_uom = "m"
    Length_uom = "m"
    Length = [
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "2",
        "3",
        "4",
        "5",
        "2",
    ]
    Length = Length[k % 14]
    Height_uom = ("m",)
    Height = [
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "1",
        "2",
        "3",
        "4",
        "5",
        "2",
    ]
    Height = Height[k % 14]
    Article_Category_Identifier = [
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87822",
        "87819,87820,87821,87823",
        "87819,87820,87821,87824",
        "87819,87820,87821,87825",
        "87819,87820,87821,87822",
    ]
    Article_Category_Identifier = Article_Category_Identifier[k % 14]
    Other_article_identifiers = [
        "item number:8570w, serial number:123123",
        "item number:8570w, serial number:123124",
        "item number:8570w, serial number:123125",
        "item number:8570w, serial number:123126",
        "item number:8570w, serial number:123127",
        "item number:8570w, serial number:123128",
        "item number:8570w, serial number:123129",
        "item number:8570w, serial number:123130",
        "item number:8570w, serial number:123131",
        "item number:8570w, serial number:123132",
        "item number:8570w, serial number:123133",
        "item number:8570w, serial number:123134",
        "item number:8570w, serial number:123135",
        "item number:8570w, serial number:123132",
    ]
    Other_article_identifiers = Other_article_identifiers[k % 14]
    Primary_article_identifier_value = random.sample(range(1, 20000), 1)[0]
    return [
        mfr,  # mfr if Levels[k + 1][1] == "No" else None,
        mpn
        if mpn
        else Primary_article_identifier_value,  # Primary_article_identifier_value if Levels[k + 1][1] == "No" else None,
        Levels[k][0],
        Number_of_Units_quantity,
        Levels[k + 1][1]
        if k + 2 != max_rows
        else "No",  # is_complex_object if k + 2 != max_rows else "No",  # Levels[k + 1][1],
        Artilce_Name,
        *Other_article_Names,
        *Primary_article_identifier_type,
        mpn if mpn else Primary_article_identifier_value,
        Other_article_identifiers,
        Article_Category_Identifier,
        *Production_in_European_Union,
        Height,
        *Height_uom,
        Length,
        *Length_uom,
        Width,
        *Width_uom,
        Diameter,
        *Diameter_uom,
        Density,
        *Density_uom,
        Weight,
        *Weight_uom,
        Volume,
        *Volume_uom,
        Colour,
        Other_Characteristics,
        Safe_use_instructions,
        Substance if Levels[k + 1][1] == "No" else None,
        CAS if Levels[k + 1][1] == "No" else None,
        *Concentration_Range,
        Material_Category,
        Additional_material_characteristics,
        None,
        *Candidate_list_substance_no_longer_present,
    ]


Levels = None


def create_manufacturer(
    args: Namespace,
    output_fname: str,
) -> dict:
    """
    Create manufacturer recrods for all the unique manufacturer given
    in master component file
    Args:
      args(dictionary):
        input : Input master component file name ( with path )
        sheet : Sheet name , containing input data ( if not given first sheet )
        header : Header row number. Data assumed to one row after header till
                  end of the file
    Returns:
     Dictionay
      status: 'Success' or 'Failure'
      msg:     If failure will contain error message else empty string
    """
    global Levels
    Levels = create_levels(args.number)

    result = {
        "status": "Success",
        "msg": "",
    }

    try:
        # ~ If we set read only then hyper link is not read
        ip_workbook = xl.load_workbook(filename=args.iref, data_only=True)
        if args.sheet1:
            ip_sheet1 = ip_workbook[args.sheet1]

        else:
            ip_sheet1 = ip_workbook["Substances from ATOM tool"]

        if not args.number:
            args.number = ip_sheet1.max_row
        # print([cell.value for cell in ip_sheet1[1]])
    except Exception as err:
        result["status"] = "Failure1"
        result["msg"] = str(err)
        return result

    Substance_column = find_Substance_column(ip_sheet1, args.header)
    if Substance_column < 0:
        result["status"] = "Failure2"
        result["msg"] = "Substance column not present in input file"
        return result

    CAS_Number_column = find_CAS_Number_column(ip_sheet1, args.header)
    if CAS_Number_column < 0:
        result["status"] = "Failure3"
        result["msg"] = "CAS Number is not present in input file"
        return result

    try:
        # ~ If we set read only then hyper link is not read
        ip_workbook = xl.load_workbook(filename=args.input, data_only=True)
        if args.sheet:
            ip_sheet = ip_workbook[args.sheet]
        else:
            ip_sheet = ip_workbook.active
    except Exception as err:
        result["status"] = "Failure4"
        result["msg"] = str(err)
        return result

    mfr_column = find_mfr_column(ip_sheet, args.header)
    if mfr_column < 0:
        result["status"] = "Failure5"
        result["msg"] = "Manufacturer Name column not present in input file"
        return result

    mpn_column = find_mpn_column(ip_sheet, args.header)
    if mpn_column < 0:
        result["status"] = "Failure6"
        result[
            "msg"
        ] = "Manufacturer part number column not present in input file"
        return result

    mfr_names = get_mfr_names(ip_sheet, args.header, mfr_column, args.number)
    mpn_names = get_mpn_names(ip_sheet, args.header, mpn_column, args.number)
    Substance_names = get_Substance_Name(
        ip_sheet1, args.header, Substance_column, args.number
    )
    CAS_Names = get_CAS_Number(
        ip_sheet1, args.header, CAS_Number_column, args.number
    )

    try:
        # ~ If we set read only then hyper link is not read
        ip_workbook = xl.load_workbook(filename=args.ibom)
        ip_sheet2 = ip_workbook.active
    except Exception as err:
        result["status"] = "Failure--"
        result["msg"] = str(err)
        return result
    ibom_mfr_column = find_ibom_mfr_column(ip_sheet2, args.header)
    ibom_mpn_column = find_ibom_mpn_column(ip_sheet2, args.header)
    ibom_mfr_names = get_mfr_names(
        ip_sheet2, args.header, ibom_mfr_column, args.number
    )
    ibom_mpn_names = get_mpn_names(
        ip_sheet2, args.header, ibom_mpn_column, args.number
    )

    for i, ibom_mfr_name in enumerate(ibom_mfr_names):
        mfr_names[i] = ibom_mfr_name
    for i, ibom_mpn_name in enumerate(ibom_mpn_names):
        mpn_names[i] = ibom_mpn_name

    master_workbook = xl.Workbook()
    master_worksheet = master_workbook.active
    master_worksheet.title = "DATA"

    try:
        ## First append header row
        master_worksheet.append(column_names)
        for mfr, mpn, Substance, CAS in zip(
            mfr_names, mpn_names, Substance_names, CAS_Names
        ):
            mfr_record = get_mfr_record(mfr, mpn, Substance, CAS, args.number)
            master_worksheet.append(mfr_record)

    except Exception as err:
        result["status"] = "Failure7"
        print(traceback.format_exc())
        result["msg"] = str(err)
        return result

    if args.validation:
        for validation_no in args.validation:
            if validation_no == 1:
                QLT707(master_workbook, args.header)
            elif validation_no == 2:
                QLT708(master_workbook, args.header)
            elif validation_no == 3:
                QLT709(master_workbook, args.header)
            elif validation_no == 4:
                QLT710(master_workbook, args.header)
            elif validation_no == 5:
                QLT711(master_workbook, args.header)
            elif validation_no == 6:
                QLT712(master_workbook, args.header)
            elif validation_no == 7:
                QLT713(master_workbook, args.header)
            elif validation_no == 8:
                QLT704(master_workbook, args.header)
            elif validation_no == 9:
                QLT703(master_workbook, args.header)
            elif validation_no == 10:
                BR705(master_workbook, args.header)

    master_workbook.save(output_fname)

    return result


def form_fname(dir_name: str, fname_base: str, fname_ext: str) -> str:
    """
    form output filename from the given parameters
    Args:
      dir_name(str): Output directory or file name
      fname_base(str): If output directory given, then it should give output
        file's base name
      fname_ext(str): If output directory given, then it should give file's
        extension
    Returns:
      str: Output file name, if input is file name, then it is return as it is
        if the input is directory name, then output will be of the form
        <dir_name>/<fname_base>_<timestamp>.<fname_ext>
    """

    dir_path = Path(dir_name)

    ## If not directory then assume it is file name
    if not dir_path.is_dir():
        return dir_name

    dir_path = Path.absolute(dir_path)
    cur_time = int(datetime.timestamp(datetime.now()))
    output_fname = fname_base + "_" + str(cur_time) + "." + fname_ext
    output_fname = Path.joinpath(dir_path, output_fname)
    return output_fname


def args_check(args: Namespace) -> dict:
    """'
    check input arguments
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    ip_file = Path(args.input)
    if not ip_file.is_file():
        result["status"] = "Failure8"
        result["msg"] = f"Input is not a valid file: {args.input}"
        return result

    ip_file = Path(args.iref)
    if not ip_file.is_file():
        result["status"] = "Failure9"
        result["msg"] = f"Input is not a valid file: {args.iref}"
        return result

    ip_file = Path(args.ibom)
    if not ip_file.is_file():
        result["status"] = "Failure--"
        result["msg"] = f"Input is not a valid file: {args.ibom}"
        return result

    op_path = Path(args.output)
    ## Output is either directory or file
    if (not op_path.is_dir()) and (not op_path.parent.is_dir()):
        result["status"] = "Failure10"
        result["msg"] = f"Output path is not a valid path: {args.output}"
        return result

    if args.header < 1:
        result["status"] = "Failure11"
        result["msg"] = f"Header Row cannot be less than 1: {args.header}"
        return result

    if args.number < 1:
        result["status"] = "Failure12"
        tmp = "Number of Rows to extract cannot be less than 1: "
        result["msg"] = f"{tmp}: {args.number}"
        return result

    if args.validation and all(val < 1 for val in args.validation):
        result["status"] = "Failure12"
        tmp = "The Validation number cannot be less than 1: "
        result["msg"] = f"{tmp}: {args.validation}"
        return result

    return result


def create_parser() -> Namespace:

    description = """ 
      Given input mastercomponent data file and a reference file that contains the cas
      number and substance names extract unique manufacturer's
      name and create manufacturers records for them as per template file
      Note: For output option if *directory* name is given, then  generated 
      file will be saved as 'Scip_Artilcedata_<timestamp>.xlsx in that directory. 
      If *filename" is given, then output will be saved in that name ( please
      provide proper path for the output )
      Validation Rules and Assigned number:
      QLT707 - 1 
      QLT708 - 2
      QLT709 - 3
      QLT710 - 4
      QLT711 - 5
      QLT712 - 6
      QLT713 - 7
      QLT704 - 8
      QLT703 - 9
      BR705  - 10
      


    """

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "--input",
        "-i",
        type=str,
        required=True,
        help="Input master component file",
    )
    parser.add_argument(
        "--sheet",
        "-s",
        type=str,
        required=False,
        help="Input Sheet Name, by default first sheet data will be taken",
    )
    parser.add_argument(
        "--iref",
        "-ir",
        type=str,
        required=True,
        help="Input Article example component file",
    )
    parser.add_argument(
        "--ibom",
        "-ib",
        type=str,
        required=True,
        help="Input Bom",
    )
    parser.add_argument(
        "--sheet1",
        "-s1",
        type=str,
        required=False,
        # help="Input Sheet Name, by default first sheet data will be taken",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        required=False,
        default="../tmp",
        help="Path to destination directory or filename",
    )
    parser.add_argument(
        "--number",
        "-n",
        type=int,
        required=False,
        help="Number of Records to extract",
        default=1,
    )
    parser.add_argument(
        "--validation",
        "-v",
        type=int,
        nargs="+",
        required=False,
        help="To get Required Validation",
    )
    parser.add_argument(
        "--header",
        "-r",
        type=int,
        required=False,
        help="Input file Header Row Number",
        default=1,
    )
    return parser.parse_args()


def find_column_no_by_column_name(
    column_name: str, ip_sheet: xl.worksheet, hdr_row_no: int
) -> int:
    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index(column_name)
    except ValueError:
        column_number = -1
    return column_number


def QLT707(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT707"

    col_1 = find_column_no_by_column_name(
        "height", master_worksheet, header_no
    )
    col_2 = find_column_no_by_column_name(
        "heightuom", master_worksheet, header_no
    )

    col_nos = [col_1, col_2]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT708(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT708"

    col_1 = find_column_no_by_column_name(
        "length", master_worksheet, header_no
    )
    col_2 = find_column_no_by_column_name(
        "lengthuom", master_worksheet, header_no
    )

    col_nos = [col_1, col_2]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT709(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT709"

    col_1 = find_column_no_by_column_name("width", master_worksheet, header_no)
    col_2 = find_column_no_by_column_name(
        "widthuom", master_worksheet, header_no
    )

    col_nos = [col_1, col_2]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT710(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT710"

    col_1 = find_column_no_by_column_name(
        "diameter", master_worksheet, header_no
    )
    col_2 = find_column_no_by_column_name(
        "diameteruom", master_worksheet, header_no
    )

    col_nos = [col_1, col_2]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT711(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT711"

    col_1 = find_column_no_by_column_name(
        "density", master_worksheet, header_no
    )
    col_2 = find_column_no_by_column_name(
        "densityuom", master_worksheet, header_no
    )

    col_nos = [col_1, col_2]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT712(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT712"

    col_1 = find_column_no_by_column_name(
        "weight", master_worksheet, header_no
    )
    col_2 = find_column_no_by_column_name(
        "weightuom", master_worksheet, header_no
    )

    col_nos = [col_1, col_2]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT713(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT713"

    col_1 = find_column_no_by_column_name(
        "volume", master_worksheet, header_no
    )
    col_2 = find_column_no_by_column_name(
        "volumeuom", master_worksheet, header_no
    )

    col_nos = [col_1, col_2]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT704(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT704"

    col_1 = find_column_no_by_column_name(
        "otherarticlenames", master_worksheet, header_no
    )

    col_nos = [col_1]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def QLT703(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "QLT703"

    col_1 = find_column_no_by_column_name(
        "otherarticleidentifiers", master_worksheet, header_no
    )

    col_nos = [col_1]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = None


def BR705(master_workbook: xl.Workbook, header_no: int) -> None:

    master_worksheet = master_workbook.copy_worksheet(master_workbook.active)
    master_worksheet.title = "BR705"

    col_1 = find_column_no_by_column_name(
        "primaryarticleidentifiertype", master_worksheet, header_no
    )

    col_nos = [col_1]
    for i, row in enumerate(master_worksheet.iter_rows()):
        if i == 0:
            continue
        for j, cell in enumerate(row):
            if j in col_nos:
                cell.value = "scip number"


if __name__ == "__main__":

    pp = pprint.PrettyPrinter(indent=2)

    ## Default output file name
    output_base = "Scip_Article_Data"
    output_ext = "xlsx"

    args = create_parser()
    result = args_check(args)
    if result["status"] == "Success":
        output_fname = form_fname(args.output, output_base, output_ext)
        args.number += 1
        result = create_manufacturer(args, output_fname)

    print("Result Status: " + result["status"])
    if result["status"] == "Success":
        print("Output file is saved in " + str(output_fname))
    else:
        print("Error Message: " + str(result["msg"]))
