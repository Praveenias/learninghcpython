import os
import openpyxl
from openpyxl import load_workbook
import argparse
import pathlib
status = {}

## Command Line Processing
cli_parser = argparse.ArgumentParser(
    prog='split_excel',
    description='Split given work sheet into multiple files')

# Add the arguments
cli_parser.add_argument('--input', '-i', type=str, required=True,
                        help='Input path folder')

args = cli_parser.parse_args()
ip_fpath = args.input


def check_input(ip_fpath):
  result = {
    'status':'Success',
    'msg'   : '',
    }
  ip_path = pathlib.Path(ip_fpath) 
  if ( not ip_path.is_dir()):
    result['status'] = 'Failure'
    result['msg'] = (f'Input is not a valid file: {ip_fpath}')
    return result


# The below method is used to read data from an active worksheet and store it in memory.
check = check_input(ip_fpath)
def reader(file):
    global path
    abs_file = os.path.join(path, file)
    wb_sheet = load_workbook(abs_file).active
    try:
        rows = []
        # min_row is set to 2, to ignore the first row which contains the headers
        for row in wb_sheet.iter_rows(min_row=2):
            row_data=[]
            for cell in row:
                row_data.append(cell.value)
            # Creating a list of lists, where each list contain a typical row's data
            rows.append(row_data)
        return rows
    except Exception as err:
        status['msg']='failure'
        print(err)

# Folder in which my source excel sheets are present
path=ip_fpath
# To get the list of excel files
files = os.listdir(path)
op_workbook=openpyxl.Workbook()
op_worksheet = op_workbook.active
if files:
    headersheet = load_workbook(os.path.join(path,files[0])).active
    headername =  [cell.value for cell in headersheet[1]]
    op_worksheet.append(headername)
try:
    for file in files:
        rows = reader(file)
        for row in rows:
            op_worksheet.append(row)
        op_workbook.save('output.xlsx')
except Exception as err:
    status['msg']='failure'
    print(err)

'''
Sample command to run this program
usage: split_excel [-h] --input INPUT
    eg: python multitoone_openpyxl.py --input data  
'''