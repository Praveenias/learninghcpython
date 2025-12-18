"""
  Given any excel file contain large number of records, it extracts given
  number of random row from that and store it is output file.
  Useful to get subset of Fmd from large set of Fmd data
"""

import pprint
from datetime import datetime
from pathlib import Path
import argparse
from argparse import Namespace
import openpyxl as xl
from openpyxl.cell import cell

## To write Excel work book
from openpyxl.workbook import workbook


def create_fmd(args: Namespace, output_fname: str) -> dict:
    """
    Create records of FMD excel file from given
    input file
    Args:
      args(dictionary):
        input : Input fmd component file name ( with path )
        sheet : Sheet name , containing input data ( if not given first sheet )
        header : Header row number. Data assumed to one row after header till
                  end of the file
        number : Number of records in output file
    Returns:
     Dictionay
      status: 'Success' or 'Failure'
      msg:     If failure will contain error message else empty string
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    try:
        # ~ If we set read only then hyper link is not read
        ip_workbook = xl.load_workbook(filename=args.input, data_only=True)
        ip_ref_workbook = xl.load_workbook(filename=args.ipref, data_only=True)
        if args.sheet:
            ip_sheet = ip_workbook[args.sheet]
        else:
            ip_sheet = ip_workbook.active
        ip_ref_sheet = ip_ref_workbook.active
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result

    population = range((args.header + 1), ip_sheet.max_row)
    if args.number >= len(population):
        result["status"] = "Failure"
        tmp = "Output Number of rows is greater than number of rows in "
        tmp = tmp + " input file. "
        tmp1 = "Input no.of rows: {ip_sheet.max_rows}"
        tmp2 = "Output no.of rows: {args.number}"
        result["msg"] = f"{tmp} {tmp1} {tmp2}"
        return result

    row_numbers = range(2, args.number + 1)

    fmd_workbook = xl.Workbook()
    fmd_worksheet = fmd_workbook.active
    fmd_worksheet.title = "Data"

    j = 2
    try:
        ## First append header row
        fmd_worksheet.append(cell.value for cell in ip_sheet[args.header])
        for i in row_numbers:
            cell_list = [cell.value for cell in ip_sheet[i]]
            if cell_list[0]:
                cell_list[0] = ip_ref_sheet[j][0].value
                cell_list[1] = ip_ref_sheet[j][1].value
                j += 1
            fmd_worksheet.append(cell_list)
        fmd_workbook.save(output_fname)
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result

    return result


def form_fname(dir_name: str, fname_base: str, fname_ext: str) -> str:
    """
    form output filename from the given parameters
    Args:
      dir_name(str): Output directory or file name
      fname_base(str): If output directory given, then it should give output
        file's base name
      fname_ext(str): If output directory given, then it should give file's
        extension
    Returns:
      str: Output file name, if input is file name, then it is return as it is
        if the input is directory name, then output will be of the form
        <dir_name>/<fname_base>_<timestamp>.<fname_ext>
    """

    dir_path = Path(dir_name)

    ## If not directory then assume it is file name
    if not dir_path.is_dir():
        return dir_name

    dir_path = Path.absolute(dir_path)
    cur_time = int(datetime.timestamp(datetime.now()))
    output_fname = fname_base + "_" + str(cur_time) + "." + fname_ext
    output_fname = Path.joinpath(dir_path, output_fname)
    return output_fname


def args_check(args: Namespace) -> dict:
    """'
    check input arguments
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    ip_file = Path(args.input)
    if not ip_file.is_file():
        result["status"] = "Failure"
        result["msg"] = f"Input is not a valid file: {args.input}"
        return result

    op_path = Path(args.output)
    ## Output is either directory or file
    if (not op_path.is_dir()) and (not op_path.parent.is_dir()):
        result["status"] = "Failure"
        result["msg"] = f"Output path is not a valid path: {args.output}"
        return result

    if args.header < 1:
        result["status"] = "Failure"
        result["msg"] = f"Header Row cannot be less than 1: {args.header}"
        return result

    if args.number < 1:
        result["status"] = "Failure"
        tmp = "Number of Rows to extract cannot be less than 1: "
        result["msg"] = f"{tmp}: {args.number}"
        return result
    return result


def create_parser() -> Namespace:

    description = """ 
      Given input excel data file, extract given number of random rows 
      and store the same in output file. Note: For output option if *directory* 
      name is given, then  generated file will be saved as 
      'mastercomp_random_<timestamp>.xlsx in that directory.  If *filename" 
      is given, then output will be saved in that name ( please provide proper
      path for the output )
    """

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "--input", "-i", type=str, required=True, help="Input excel file"
    )
    parser.add_argument(
        "--ipref", "-ir", type=str, required=True, help="Input ref excel file"
    )
    parser.add_argument(
        "--sheet",
        "-s",
        type=str,
        required=False,
        help="Input Sheet Name, by default first sheet data will be taken",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        required=False,
        default="../tmp",
        help="Path to destination directory or filename",
    )
    parser.add_argument(
        "--number",
        "-n",
        type=int,
        required=False,
        help="Number of Records to extract",
        default=1,
    )
    parser.add_argument(
        "--header",
        "-r",
        type=int,
        required=False,
        help="Header Row Number. Note: Excel row number starts with 1",
        default=1,
    )
    return parser.parse_args()


if __name__ == "__main__":

    pp = pprint.PrettyPrinter(indent=2)

    ## Default output file name
    output_base = "fmd"
    output_ext = "xlsx"

    args = create_parser()
    result = args_check(args)
    if result["status"] == "Success":
        output_fname = form_fname(args.output, output_base, output_ext)
        result = create_fmd(args, output_fname)

    print("Result Status: " + result["status"])
    if result["status"] == "Success":
        print("Output file is saved as " + str(output_fname))
    else:
        print("Error Message: " + result["msg"])
