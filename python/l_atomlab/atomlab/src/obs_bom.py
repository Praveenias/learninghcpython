"""
  Given any excel file contain mastercomponents,
  it creates BOM file as required
  by the ATOM template
"""

import pprint
from datetime import datetime
from pathlib import Path
import argparse
from argparse import Namespace
import re
from faker import Faker
import random

import openpyxl as xl

## To write Excel work book
from openpyxl.workbook import workbook

column_names = [
    "Customer Part number",
    "Customer part number description",
    "Supplier Name",
    "Manufacturer name",
    "Imported Manufacturer Name",
    "Manufacturer part number",
    "Item Type",
    "Part Make",
    "Qty",
    "UOM",
]


suffix_l = ["inc", "corp", "corporation", "llc", "ltd", "limited"]


def normalize_string(raw_string):
    """
    Trim the string
    Convert all non-Alphanumeric characters to _
    Convert to lowercase
    If the last word is any one of suffix_l then remove that ( useful
     to normalize manufacturer name )
    Reduce consequtive _ characters into single _
    """
    if not raw_string:
        return None
    tmp_s = str(raw_string).strip()
    tmp_s = re.sub("[^\w]+", " ", tmp_s)
    tmp_l = tmp_s.lower().split()
    if tmp_l[-1] in suffix_l:
        del tmp_l[-1]
    res = "_".join(tmp_l)
    return res


def find_mfr_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("manufacturername")
    except ValueError:
        column_number = -1
    return column_number


def find_mpn_column(ip_sheet: xl.worksheet, hdr_row_no: int) -> int:
    """
    Note: It return column number starts with zero
    """

    pat = re.compile(r"[\s_-]")
    # ~ hdr_row = [cell.value.strip().lower() for cell in ip_sheet[hdr_row_no]]
    hdr_row = []
    for cell in ip_sheet[hdr_row_no]:
        val = cell.value.strip().lower()
        val = re.sub(pat, "", val)
        hdr_row.append(val)
    try:
        column_number = hdr_row.index("manufacturerpartnumber")
    except ValueError:
        column_number = -1
    return column_number


def get_mfr_names(ip_sheet: xl.worksheet, hdr_row_no: int, mfr_col: int) -> list:
    st_row = hdr_row_no + 1
    mfr_names = [
        ip_sheet[r][mfr_col].value for r in range(st_row, ip_sheet.max_row + 1)
    ]
    return mfr_names


def get_mpn_names(ip_sheet: xl.worksheet, hdr_row_no: int, mpn_col: int) -> list:
    st_row = hdr_row_no + 1
    mpn_names = [
        ip_sheet[r][mpn_col].value for r in range(st_row, ip_sheet.max_row + 1)
    ]
    return mpn_names


def get_mfr_record(mfr: str, mpn: str) -> list:
    """
    ToDo: Need to fill up all the mandatory column
    """
    Customer_Part_number = [
        "00062-08-09827",
        "00062-08-09828",
        "00062-08-09829",
        "00062-08-09830",
        "00062-08-09831",
        "00004-02-00025",
        "00004-02-00021",
        "00005-08092",
        "00005-02-00002",
        "00005-07-00022",
        "00004-95080",
        "00062-08-09822",
        "00062-08-09823",
        "00062-08-09824",
        "00062-08-09825",
        "00062-08-09826",
        "00015-04-00523",
        "00015-00375",
        "00015-04-00184",
        "00015-04-00523",
        "00015-04-00466",
    ]
    Customer_Part_number = random.sample(Customer_Part_number, 1)
    Customer_part_number_description = [
        "IC,TSTR",
        "CONN, SMA, RT-ANGLE, LTQ-VELOS ONLY",
        "DIODE,TVS,UNI-DIR,600W,36V,SMB,RoHS",
        "DIODE,RECT,S1GB,1A,400V,SMT-SMB,RoHS",
        "DIODE,RECT,SCHTKY,5A,40V,B540C,SMC,RoHS",
        "DIODE,RECT,SCHTKY,5A,40V,B540C,SMC,RoHS",
        "CONN-ACC,JACK,4-40 X 0.25L,STEEL,RoHS",
        "CNTCT,SKT,12AWG,STD,CRIMP,ELCON,RoHS",
        "RES,THK FILM,162K,1/4W,1%,SM1206,RoHS",
        "RES,THK-FLM 100K 1/8W 1% SMT0805,RoHS",
        "RES,THK-FLM 2.74K 1/4W 1% SMT1206,RoHS",
        "RES,Chip,1/4 W,33,5%,SMT1206,RoHS",
    ]
    Customer_part_number_description = random.sample(
        Customer_part_number_description, 1
    )
    Item_Type = "Purchased"
    Part_Make = "Generic"
    Qty = "1"
    UOM = "EA"

    return [
        *Customer_Part_number,
        *Customer_part_number_description,
        None,
        mfr,
        mfr,
        mpn,
        Item_Type,
        Part_Make,
        Qty,
        UOM,
    ]


def create_manufacturer(
    args: Namespace,
    output_fname: str,
) -> dict:
    """
    Create manufacturer recrods for all the unique manufacturer given
    in master component file
    Args:
      args(dictionary):
        input : Input master component file name ( with path )
        sheet : Sheet name , containing input data ( if not given first sheet )
        header : Header row number. Data assumed to one row after header till
                  end of the file
    Returns:
     Dictionay
      status: 'Success' or 'Failure'
      msg:     If failure will contain error message else empty string
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    try:
        # ~ If we set read only then hyper link is not read
        ip_workbook = xl.load_workbook(filename=args.input, data_only=True)
        if args.sheet:
            ip_sheet = ip_workbook[args.sheet]
        else:
            ip_sheet = ip_workbook.active
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result

    mfr_column = find_mfr_column(ip_sheet, args.header)
    if mfr_column < 0:
        result["status"] = "Failure"
        result["msg"] = "Manufacturer Name column not present in input file"
        return result

    mpn_column = find_mpn_column(ip_sheet, args.header)
    if mpn_column < 0:
        result["status"] = "Failure"
        result["msg"] = "Manufacturer part number column not present in input file"
        return result

    mfr_names = get_mfr_names(ip_sheet, args.header, mfr_column)
    mpn_names = get_mpn_names(ip_sheet, args.header, mpn_column)

    master_workbook = xl.Workbook()
    master_worksheet = master_workbook.active
    master_worksheet.title = "BOM"

    try:
        ## First append header row
        master_worksheet.append(column_names)
        for mfr, mpn in zip(mfr_names, mpn_names):
            mfr_record = get_mfr_record(mfr, mpn)
            master_worksheet.append(mfr_record)
        master_workbook.save(output_fname)
    except Exception as err:
        result["status"] = "Failure"
        result["msg"] = str(err)
        return result
    return result


def form_fname(dir_name: str, fname_base: str, fname_ext: str) -> str:
    """
    form output filename from the given parameters
    Args:
      dir_name(str): Output directory or file name
      fname_base(str): If output directory given, then it should give output
        file's base name
      fname_ext(str): If output directory given, then it should give file's
        extension
    Returns:
      str: Output file name, if input is file name, then it is return as it is
        if the input is directory name, then output will be of the form
        <dir_name>/<fname_base>_<timestamp>.<fname_ext>
    """

    dir_path = Path(dir_name)

    ## If not directory then assume it is file name
    if not dir_path.is_dir():
        return dir_name

    dir_path = Path.absolute(dir_path)
    cur_time = int(datetime.timestamp(datetime.now()))
    output_fname = fname_base + "_" + str(cur_time) + "." + fname_ext
    output_fname = Path.joinpath(dir_path, output_fname)
    return output_fname


def args_check(args: Namespace) -> dict:
    """'
    check input arguments
    """

    result = {
        "status": "Success",
        "msg": "",
    }

    ip_file = Path(args.input)
    if not ip_file.is_file():
        result["status"] = "Failure"
        result["msg"] = f"Input is not a valid file: {args.input}"
        return result

    op_path = Path(args.output)
    ## Output is either directory or file
    if (not op_path.is_dir()) and (not op_path.parent.is_dir()):
        result["status"] = "Failure"
        result["msg"] = f"Output path is not a valid path: {args.output}"
        return result

    if args.header < 1:
        result["status"] = "Failure"
        result["msg"] = f"Header Row cannot be less than 1: {args.header}"
        return result

    return result


def create_parser() -> Namespace:

    description = """ 
      Given input mastercomponent data file, extract unique manufacturer's
      name and create manufacturers records for them as per template file
      Note: For output option if *directory* name is given, then  generated 
      file will be saved as 'manufacturers_<timestamp>.xlsx in that directory. 
      If *filename" is given, then output will be saved in that name ( please
      provide proper path for the output )
    """

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "--input", "-i", type=str, required=True, help="Input master component file"
    )
    parser.add_argument(
        "--sheet",
        "-s",
        type=str,
        required=False,
        help="Input Sheet Name, by default first sheet data will be taken",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        required=False,
        default="../tmp",
        help="Path to destination directory or filename",
    )
    parser.add_argument(
        "--header",
        "-r",
        type=int,
        required=False,
        help="Input file Header Row Number",
        default=1,
    )
    return parser.parse_args()


if __name__ == "__main__":

    pp = pprint.PrettyPrinter(indent=2)

    ## Default output file name
    output_base = "BOM"
    output_ext = "xlsx"

    args = create_parser()
    result = args_check(args)
    if result["status"] == "Success":
        output_fname = form_fname(args.output, output_base, output_ext)
        result = create_manufacturer(args, output_fname)

    print("Result Status: " + result["status"])
    if result["status"] == "Success":
        print("Output file is saved as " + str(output_fname))
    else:
        print("Error Message: " + str(result["msg"]))
