import openpyxl
import pathlib
import logging

Log_Format = "%(levelname)s %(asctime)s - %(message)s"




class Helper:

    def __init__(self):
        self.result={}
        p = pathlib.Path("../Output/")
        p.mkdir(parents=True, exist_ok=True)
        logging.basicConfig(filename = "../logs/logfile2.log",
                    filemode = "w",
                    format = Log_Format,
                    level=logging.INFO)
        self.logger = logging.getLogger()
        self.logger.info('started')
     

    def check_ip(self,ipfilename):
        if not pathlib.Path('../data/'+ipfilename).is_file():
            self.result['status']='Failure'
            self.result['Info'] = 'File not exists'
            return False
        return True

    def split(self,ipdict):
        constant = ipdict['constantColumn']
        repeated = ipdict['repeatedColumn']

        opwb=openpyxl.Workbook()
        opws = opwb.active  
        try:
            ipworkbook = openpyxl.load_workbook('../data/'+ipdict['input_file_name'])
            ipworksheet = ipworkbook.active

            no_of_rows = ipworksheet.max_row
            no_of_columns = ipworksheet.max_column

            self.logger.info(f'NO OF ROWS : {no_of_rows} ,NO OF COLUMNS : {no_of_columns}')
            self.logger.info(f'Output Rows in Excel file(Including HeaderName): {(no_of_rows-1)*len(repeated)+1}')


            cnstclmName = ipworksheet[f'{constant[0]}{1}':f'{constant[1]}{1}'][0]
            opws.append([cell.value for cell in cnstclmName])
            

            for data in ipworksheet.iter_rows(min_row=2,max_col=no_of_columns, max_row=no_of_rows):              
                column=data[0].row
                #self.logger.info(f'{column} nd Row completed')
                for repeat in range(len(repeated)):
                    data = ipworksheet[f'{constant[0]}{column}':f'{constant[1]}{column}'][0] +  ipworksheet[f'{repeated[repeat][0]}{column}':f'{repeated[repeat][1]}{column}'][0]
                    opws.append(tuple([dat.value for dat in data]))
                opwb.save("../Output/"+ipdict['output_file_name'])
        except Exception as err:
            self.result['status']='Failure'
            self.result['Info'] = err
            return
        self.logger.info('File successfully splitted and saved')
        self.result['status']='Success'
        self.result['msg'] = 'File successfully splitted and saved'
        self.result['Info'] = 'check log file In Log folder'

    def index(self,ipdict):
        if self.check_ip(ipdict['input_file_name']):
            self.split(ipdict)
        return self.result

    def check_partstatus(ip_str:str)->str:
        if ip_str.find("act") != -1:
            return "active"
        if ip_str.find("eol") != -1:
            return "eol"
        if ip_str.find("nrn") != -1:
            return "nrnd"
        if ip_str.find("ltb") != -1:
            return "ltb"
        if ip_str.find("appli") != -1:
            return "not applicable"
        if ip_str.find("avail") != -1:
            return "not available"   
        else:
            return ip_str

    def check_partcategory(ip_str:str)->str:
        if ip_str.find("electri") != -1:
            return "electrical & electronics"
        if ip_str.find("electro") != -1:
            return "electro-mechanical"
        if ip_str.find("mechanic") != -1:
            return "mechanical"
        if ip_str.find("appli") != -1:
            return "not applicable"
        if ip_str.find("avail") != -1:
            return "not available"   
        else:
            return ip_str

if __name__ == '__main__':
    constantColumn = ['A','B']
    repeatedColumn = [['C','F'],['G','I']]
    input_file_name = 'test.xlsx'
    output_file_name = 'output.xlsx'

   

    IpDict = {'constantColumn':constantColumn,'repeatedColumn':repeatedColumn,
                'input_file_name':input_file_name,'output_file_name':output_file_name }

    splitrow = Helper()
    result = splitrow.index(IpDict)
    print(result)
