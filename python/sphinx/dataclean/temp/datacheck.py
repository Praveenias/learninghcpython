'''
  Datacheck python file
  checking the data for given excel file is correct or incorrect
  this code will precheck the data before uploading into the tool.
  Data validation done and incorrect msges are written in issue column

'''


import json
import openpyxl
import argparse
import re
import validators
import logging
from pathlib import Path
from datetime import datetime
#from help import Helper


def gen_outfpath(out_filename:str)->Path:
  """create output file name  

  Parameters
  ----------

  out_filename : str
    output file name by the input

  Returns
  -------
  path
    out file path with outfile name
  """

  if not out_filename:
    cur_time = int(datetime.timestamp(datetime.now()))
    out_filename = "result_"+str(cur_time)
  out_dir = Path(Path.joinpath(BASE_DIR,"../output/")).mkdir(parents=True, exist_ok=True)
  dir_path = Path(Path.joinpath(BASE_DIR,"../output/"))
  
  filename = out_filename+".xlsx"
  output_fpath = Path.joinpath(dir_path, filename)
  return output_fpath

def data_type_check(dataType:str , data:str , clmName:str)->str:
  """validates the indivual row value  

  Parameters
  ----------

  dataType : str
    data type for that particular column
  data : str
    column data
  clmName : str
    column name

  Returns
  -------
  str
    output string 
  """
  
  if dataType == 'int' and not (str(data).isdigit() and data):
    return clmName + " must be postive integer;"
  if dataType == 'intfloat' and not isinstance(data,(int,float)) :
    return clmName +" is invalid amount; "
  if dataType == 'float' and not isinstance(data,float) :
    return clmName +" is invalid ; "
  if dataType == 'date':
    try:
      
      if ip_args.datatype == 'regulation':
        datetime.strptime(data, "%d-%b-%y")
      elif ip_args.datatype == 'coc':
        datetime.strptime(data, '%Y-%m-%d') #coc date format YYYY-MM-DD
      else:
        datetime.strptime(data,"%Y-%m-%d") #dd-mm-yyyy
    except Exception as e:
      return clmName +" is invalid date; "
  if dataType == 'url':
    if not validators.url(data):
      return clmName +" is invalid url ; "
  if dataType == 'email':
    #regex = re.compile(r'([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@([A-Za-z0-9]+[.-_])+(\.[A-Z|a-z]{2,})+')
    if not validators.email(data):
      return clmName +" is invalid email ; " #check spaces also
  if dataType == 'capro65V':
    regex = r'ca prop 65-[0-9]{4}'
    if not re.fullmatch(regex,data.lower()):
      return clmName +" is invalid format ; "
  if dataType == 'length250':
    if len(data)>250:
      return clmName +" should be only 250 charecters ; ",
  if dataType == 'f_date':
    today_date=datetime.today()
    if today_date > data:
      return clmName+" should not be past date"
  return ""

def normalize_string(raw_string:str,remove_suffix:bool=False)->str:
  """normalize the given string 

  Parameters
  ----------

  raw_string : str
    input string
  remove_suffix : bool
    suffix word

  Returns
  -------
  str
    output normalized string 
  """

  suffixL = ['inc', 'corp', 'corporation', 'llc', 'ltd', 'limited']
  if (not raw_string):
      return None
  tmp_s = str(raw_string).strip()
  tmp_s = re.sub('[^\w]+', ' ',tmp_s)
  tmp_l = tmp_s.lower().split()
  if (tmp_l[-1] in suffixL) and remove_suffix: del tmp_l[-1]
  res = '_'.join(tmp_l)
  return res
  
def check_data(row_data:dict,template_data:dict,is_fmd:bool) -> str:
  """Validates the single row and return the issue  

  Parameters
  ----------

  row_data : dict
    single row data
  template_data : dict
    config template data

  Returns
  -------
  str
    output string 
  """

  issue = ''
  column_list = template_data.keys()
  try:      
    for clm_name, data in row_data.items():
      if isinstance(data, str) and data.startswith("="):
        issue += f'may be formula used in {clm_name} ;'
      if (not clm_name in column_list): 
        continue 
      column_template = template_data[clm_name]
      column_fields = column_template.keys()    
      if 'mandatory' in column_fields:
        if is_fmd and data is None:
          continue
        if (column_template['mandatory']) and (data is None):
          issue += f' {clm_name}: Mandatory column, but data not present; '
          continue
        if (not column_template['mandatory']) and (data is None):
          continue
      if 'options' in column_fields:
        data = data.lower().strip()
        if clm_name == "part_life_cycle_status":pass
          #data = Helper.check_partstatus(data)        
        if clm_name == "part_category":pass
          #data = Helper.check_partcategory(data)
        if data not in column_template['options']:
          issue += f' {clm_name}: Given {data} is not an allowed option; '
      if 'datatype' in column_fields:
        issue += data_type_check(column_template['datatype'],data,clm_name)
  except Exception as err:   
    issue  = err
  return issue

def get_fmd_issue(data_d:dict,template_data:dict,check_clm:str)->str:
  """check mandatory details for fmd file only  

  Parameters
  ----------

  data_d : dict
    list of clm name and value
  check_clm : excel work sheet
    clm that needed to check
  template_data : dict
    config template data

  Returns
  -------
  str
    output string 
  """
  issue=''
  try:
    if check_clm == 'req-field-1':
      list = template_data['req-field-1']+template_data['req-field-2']+template_data['req-field-3']
    elif check_clm == 'req-field-2':
      list = template_data['req-field-2']+template_data['req-field-3']
    else:
      list = template_data['req-field-3']
    for clm_name in list:
      if 'mandatory' in template_data[clm_name] and data_d[clm_name] == None:
        issue += f' {clm_name}: Mandatory column, but data not present; '
    return issue
  except Exception as e:
    print(e)
    return ''



def write_issue(headername:list , ws:openpyxl.worksheet,
                args :argparse.Namespace , template_data : dict) -> str:
  """extract data from the excel , check for isssue and uniqueness  

  Parameters
  ----------

  headername : list
    list of header names
  ws : excel work sheet
    input excel worksheet
  args : arguments
    input arguments
  template_data : dict
    config template data

  Returns
  -------
  str
    output string 
  """

  result = ''

  opFileName=args.outfilename
  headerRow = args.headerrow
  columnSize = ws.max_column
  rowSize = ws.max_row
  print(f'In Data file, no of rows: {rowSize} , no of columns: {columnSize}')
  is_fmd = True if args.datatype == 'fmd' else False

  try:
    opWorkBook = openpyxl.Workbook()
    opWorkSheet = opWorkBook.active

    opWorkSheet.append(headername+['issue','duplicate'])

  except Exception as err:
    result = f'Opening output Excel file error: #{err}'
    return result
  
  
  result = ''
  unique_cols=[]
  if ('unique' in template_data):
    unique_cols = template_data.pop('unique')
  unique_str_d = { }
  try:          
    row_number = headerRow
    for data in ws.iter_rows(min_row=headerRow+1,max_col=len(headername),
                                     max_row=rowSize,values_only=True ):
      issue=''
      row_number += 1
      emptyRow = all(ele==None for ele in data) 
      if emptyRow:continue
      data_d = dict(zip(headername,data))
      if is_fmd:
        clm_data_value = [True for clm_name in template_data['req-field-1'] if data_d[clm_name]!=None ]
        if True in clm_data_value:
          issue = get_fmd_issue(data_d,template_data,"req-field-1")         
        else :
          clm_data_value = [True for clm_name in template_data['req-field-2'] if data_d[clm_name]!=None ]
          if True in clm_data_value:
            issue = get_fmd_issue(data_d,template_data,"req-field-2")
          else:
            clm_data_value = [True for clm_name in template_data['req-field-3'] if data_d[clm_name]!=None ]
            issue = get_fmd_issue(data_d,template_data,"req-field-3")

      issue += check_data(data_d,template_data,is_fmd)

      duplicate = ''
      if (unique_cols):
        unique_s = ''
        for col in unique_cols:
          if 'normalize' in template_data[col].keys():
            unique_s += normalize_string(str(data_d[col]),True)
          else:
            unique_s += str(data_d[col])
        
        if (unique_s in unique_str_d):
          duplicate = f'Duplicate of row: {unique_str_d[unique_s]}'
        else:
          if not unique_s =='None':
            unique_str_d[unique_s] = row_number 
      
      rowValue  = list(data)+[issue,duplicate]
      opWorkSheet.append(rowValue)
    out_filename = gen_outfpath(opFileName)
    opWorkBook.save(filename= out_filename)

  except Exception as err:
    logging.warning(f'Excel Processing error: {err}')
    result = f'Excel Processing error: {err}'

  return result

def get_mandatory_columns(template_data:dict) -> list:
  """get the mandatorycolumn from the template  

  Parameters
  ----------

  template_data : dict
    config template data for repective file type

  Returns
  -------
  list
    list of mandatory columns 
  """

  result = []
  for col_name, constr_d in template_data.items():
    if (col_name == 'unique'): continue
    if ( ('mandatory' in constr_d) and (constr_d['mandatory']) ):
      result.append(col_name)
  return result

def check_main(args:argparse.Namespace , template_data:dict) -> str:
  """checks the mandatory header columns    

  Parameters
  ----------
  args : argparse.Namespace
    arguments

  template_data : dict
    config template data for repective file type

  Returns
  -------
  str
    output string 
  """

  inputfile = args.inputfile
  inSheetName = args.ipsheetname
  headerRow = args.headerrow
  ip_file = Path.joinpath(BASE_DIR,inputfile)
  try:
    ipWorkBook = openpyxl.load_workbook(filename= ip_file)           
    ipworksheet = ipWorkBook[inSheetName] if inSheetName else ipWorkBook.active
  except Exception as err:
    return f'Opening Input worksheet error: #{err}'

  header_names = [cell.value for cell in ipworksheet[headerRow]]
  if None in header_names:
    return f'{header_names.index(None)+1}: Column header is empty '

  mandatory_cols = get_mandatory_columns(template_data)
  normalized_header = list(map(normalize_string, header_names))
  if not all(item in normalized_header for item in mandatory_cols):
    header_missing = set(mandatory_cols)-set(normalized_header)
    return f'{header_missing}: Mandatory Column name(s) are missing'

  return write_issue(normalized_header,ipworksheet,args,template_data)

def choose_template(template_fname:str, input_datatype:str)->dict:
  """get the config data from the json templates  

  Parameters
  ----------
  template_fname : str
    json file path

  input_datatype : str
    type of the excel file

  Returns
  -------
  dict
    dict with config details
  """

  try:
    with open(Path.joinpath(BASE_DIR,template_fname),encoding='utf-8') as fd:
      json_data = json.load(fd)
      if (input_datatype in json_data):
        template_data = json_data[input_datatype]
      else:
        template_data = { }
  except Exception as e:
    logging.warning(f'Reading Template file failed: {e}')
    print(f'Reading Template file failed: {e}')
    template_data = { }
  return template_data
    
def check_input(args:argparse.Namespace) -> str:
  """checks the input that from the argsparser  

  Parameters
  ----------
  args : argparse.Namespace
    arguments

  Returns
  -------
  str
    string with corresponding details
  """
  
  msg = '';
  if not Path(Path.joinpath(BASE_DIR,args.inputfile)).is_file():
    msg  += ' Input excel file not present. '
  
  if args.headerrow<1 :
    msg  += ' Header Row cannot be less than 1. '
  
  return msg

def create_parser(allowed_datatype:list) -> argparse.Namespace:
  """creating input that needs from user through argument 
     parser

  Parameters
  ----------
  allowed_datatype : list
    The file location of the spreadsheet

  Returns
  -------
  argparser
    argument namespace
  """

  parser = argparse.ArgumentParser(
    prog='dataClean',
    description='Check the data of given excel file')

  parser.add_argument("--datatype","-d",required=True,type=str,
                      help="Input file classification",
                      choices=allowed_datatype)
  parser.add_argument("--inputfile","-i",required=True,
                      help="Input file name with full path and extension")
  parser.add_argument("--ipsheetname","-s", 
                      help="Input Sheet name(Case Sensitive)")
  parser.add_argument("--outfilename","-o",
                      help="Output File Name(without extension)")
  parser.add_argument("--headerrow","-r",type=int, help="Header Name Row",
                      default=1)
  # parser.add_argument("--filetype","-f",required=True,
  #                     help="enter File type",choices=file_type) 

  args = parser.parse_args()
  return args
    

if __name__ == '__main__':
  '''
  template_fname: 'rules json file'
  allowed_datatype: ' allowed file templates'
  allowed_filetype: 'says file belong to obs or comliance'

  '''
  
  print('Start of datacheck....')
  BASE_DIR = Path(__file__).resolve().parent
  logging.basicConfig(filename=Path.joinpath(BASE_DIR,'../logs/app.log'), filemode='w', level=logging.INFO,
                      format='%(asctime)s - %(levelname)s -%(message)s', datefmt='%d-%b-%y %H:%M:%S')
  # template_fname = {
  #   "obs":'./datacheck_obs_temp.json',
  #   "compliance":'./datacheck_com_temp.json'
  # }
  template_fname = './atom_templates.json'

  #~ Template file expect to have defined JSON data for these keys
  allowed_datatype = ['bom','contact','master','manufacturer', 'regulation',
                      'substance','fmd','coc','scip','rba','query_statement',
                      'exemption_catagory','exemption_list']
  #allowed_filetype = ["obs","compliance"]

  ip_args = create_parser(allowed_datatype)
  logging.info(f"input arguments{ip_args}")
  
  if not Path(Path.joinpath(BASE_DIR,template_fname)).is_file():
    print(f'{template_fname}: Template file to check data is not present ')
  else:   
    result = check_input(ip_args)
    if (not result):
      template_fields = choose_template(template_fname, ip_args.datatype)
      if (template_fields):
        result = check_main(ip_args,template_fields)       
        if result:
          print(f'Excel file processing failed: {result}')
        else:
          print(f'Excel file processed and result updated in output folder')

      else:
        print(f'In {template_fname} no template defined for {ip_args.datatype}')
    else:
      print(f'Input check failed: {result}')

   
  print('End of datacheck....')

#python datacheck.py --datatype bom --inputfile ../data/datacheck_bom_com.xlsx --outfilename result_bom 
#python datacheck.py --datatype master --inputfile ../data/datacheck_master_com.xlsx --outfilename result_master 
#python datacheck.py --datatype substance --inputfile ../data/datacheck_substance.xlsx --outfilename result_substance 
#python datacheck.py --datatype regulation --inputfile ../data/datacheck_regulation.xlsx --outfilename result_regulation 
#python datacheck.py --datatype manufacturer --inputfile ../data/datacheck_manufacturer.xlsx --outfilename result_manufacturer 
#python datacheck.py --datatype contact --inputfile ../data/datacheck_contact.xlsx --outfilename result_contact 
 