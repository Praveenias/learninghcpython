import json
import openpyxl
import argparse
import re
import validators
import logging
from pathlib import Path
from datetime import datetime


class Datacheck:

    def __init__(self) -> None:
        self.template_json_path = 'atom_templates.json'
        self.input_data = {}

    def check_input(self,user_input) -> str:
        result = {'status':'success','msg':''}   
        if user_input['header_row'] < 1 :
            result['status'] = 'error'
            result['msg'] = ' Header Row cannot be less than 1. '   
        return result

    def choose_template(self)->dict:

        try:
            with open(Path(self.template_json_path),encoding='utf-8') as fd:
                json_data = json.load(fd)
                if (self.input_data['file_type'] in json_data):
                    template_data = json_data[self.input_data['file_type']]
                else:
                    template_data = {}
        except Exception as e:
            print(f'Reading Template file failed: {e}')
            template_data = {}
        return template_data

    def get_mandatory_columns(self,template_data:dict) -> list:
        result = []
        for col_name, constr_d in template_data.items():
            if (col_name == 'unique'): continue
            if ( ('mandatory' in constr_d) and (constr_d['mandatory']) ):
                result.append(col_name)
        return result
    def normalize_string(self,raw_string:str,remove_suffix:bool=False)->str:
        suffixL = ['inc', 'corp', 'corporation', 'llc', 'ltd', 'limited']
        if (not raw_string):
            return None
        tmp_s = str(raw_string).strip()
        tmp_s = re.sub('[^\w]+', ' ',tmp_s)
        tmp_l = tmp_s.lower().split()
        if (tmp_l[-1] in suffixL) and remove_suffix: del tmp_l[-1]
        res = '_'.join(tmp_l)
        return res

    def check_main(self,template_data:dict) -> str:
        result = {'status':'success','msg':''}
        inputfile = self.input_data['input_file_path']
        inSheetName = self.input_data['sheet_name']
        headerRow = self.input_data['header_row']
        ip_file = Path(inputfile)
        try:
            ipWorkBook = openpyxl.load_workbook(filename= ip_file)           
            ipworksheet = ipWorkBook[inSheetName] if inSheetName else ipWorkBook.active
        except Exception as err:
            result['status'] = 'error'
            result['msg'] = f'Opening Input worksheet error: #{err}'
            return result

        header_names = [cell.value for cell in ipworksheet[headerRow]]
        if None in header_names:
            print(header_names)
            result['status'] = 'error'
            result['msg'] = f'{header_names.index(None)+1}: Column header is empty '
            return result
        mandatory_cols = self.get_mandatory_columns(template_data)
        normalized_header = list(map(self.normalize_string, header_names))
        if not all(item in normalized_header for item in mandatory_cols):
            header_missing = set(mandatory_cols)-set(normalized_header)
            result['status'] = 'error'
            result['msg'] = f'{header_missing}: Mandatory Column name(s) are missing'
            return result
        return self.write_issue(normalized_header,ipworksheet,template_data)
        return result
    
    def gen_outfpath(self)->Path:

        if not self.input_data['out_file_name'] :
            cur_time = int(datetime.timestamp(datetime.now()))
            self.input_data['out_file_name'] = "result_"+str(cur_time)
        out_dir = Path("../output/").mkdir(parents=True, exist_ok=True)
        dir_path = Path("../output/")
        
        filename = self.input_data['out_file_name']+".xlsx"
        output_fpath = Path.joinpath(dir_path, filename)
        return output_fpath

    def data_type_check(self,dataType:str,data,clmName:str)->str:

        if dataType == 'int' and not (str(data).isdigit() and data):
            return clmName + " must be postive integer;"
        if dataType == 'intfloat' and not isinstance(data,(int,float)) :
            return clmName +" is invalid amount; "
        if dataType == 'float' and not isinstance(data,float) :
            return clmName +" is invalid ; "
        if dataType == 'date':
            try:
                if self.input_data['file_type'] == 'regulation':
                    datetime.strptime(data, "%d-%b-%y")
                elif self.input_data['file_type'] == 'coc':
                    datetime.strptime(data, '%Y-%m-%d') #coc date format YYYY-MM-DD
            except Exception as e:
                return clmName +" is invalid date; "
        if dataType == 'url':
            if not validators.url(data):
                return clmName +" is invalid url ; "
        if dataType == 'email':
            #regex = re.compile(r'([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@([A-Za-z0-9]+[.-_])+(\.[A-Z|a-z]{2,})+')
            if not validators.email(data):
                return clmName +" is invalid email ; " #check spaces also
        if dataType == 'capro65V':
            regex = r'ca prop 65-[0-9]{4}'
            if not re.fullmatch(regex,data.lower()):
                return clmName +" is invalid format ; "
        if dataType == 'length250':
            if len(data)>250:
                return clmName +" should be only 250 charecters ; ",
        if dataType == 'f_date':
            today_date=datetime.today()
            if today_date > data:
                return clmName+" should not be past date"
        return ""

    def check_data(self,row_data:dict,template_data:dict) -> dict:
        issue = ''
        column_list = template_data.keys()
        try:      
            for clm_name, data in row_data.items():
                if isinstance(data, str) and data.startswith("="):
                    issue += f'may be formula used in {clm_name} ;'
                if (not clm_name in column_list): 
                    continue 
                column_template = template_data[clm_name]
                column_fields = column_template.keys()
                if 'mandatory' in column_fields:
                    if (column_template['mandatory']) and (data is None):
                        issue += f' {clm_name}: Mandatory column, but data not present; '
                        continue
                    if (not column_template['mandatory']) and (data is None):
                        continue
                if 'options' in column_fields:
                    data = data.lower().strip()
                    # if clm_name == "part_life_cycle_status":
                    #     data = Helper.check_partstatus(data)        
                    # if clm_name == "part_category":
                    #     data = Helper.check_partcategory(data)
                    if data not in column_template['options']:
                        issue += f' {clm_name}: Given {data} is not an allowed option; '
                if 'datatype' in column_fields:
                    issue += self.data_type_check(column_template['datatype'],data,clm_name)
        except Exception as err:
            issue  = err
        return issue

    def write_issue(self,headername:list , ws:openpyxl.worksheet,
                    template_data : dict) -> str:
        result = {'status':'success','msg':''}

        opFileName=self.input_data['out_file_name']
        headerRow = self.input_data['header_row']
        columnSize = ws.max_column
        rowSize = ws.max_row
        print(f'In Data file, no of rows: {rowSize} , no of columns: {columnSize}')

        try:
            opWorkBook = openpyxl.Workbook()
            opWorkSheet = opWorkBook.active

            opWorkSheet.append(headername+['issue','duplicate'])

        except Exception as err:
            result['status'] = 'error'
            result['msg'] = f'Opening output Excel file error: #{err}'
            return result
               
        unique_cols=[]
        if ('unique' in template_data):
            unique_cols = template_data.pop('unique')
        unique_str_d = { }
        try:          
            row_number = headerRow
            for data in ws.iter_rows(min_row=headerRow+1,max_col=len(headername),
                                            max_row=rowSize,values_only=True ):

                row_number += 1
                emptyRow = all(ele==None for ele in data) 
                if emptyRow:continue

                data_d = dict(zip(headername,data))
                issue = self.check_data(data_d,template_data)

                duplicate = ''
                if (unique_cols):
                    unique_s = ''
                    for col in unique_cols:
                        if 'normalize' in template_data[col].keys():
                            unique_s += self.normalize_string(str(data_d[col]),True)
                        else:
                            unique_s += str(data_d[col])
                    
                    if (unique_s in unique_str_d):
                        duplicate = f'Duplicate of row: {unique_str_d[unique_s]}'
                    else:
                        if not unique_s =='None':
                            unique_str_d[unique_s] = row_number 
                rowValue  = list(data)+[issue,duplicate]
                opWorkSheet.append(rowValue)
            out_filename = self.gen_outfpath()
            opWorkBook.save(filename= out_filename)

        except Exception as err:
            logging.warning(f'Excel Processing error: {err}')
            result['status'] = 'error'
            result['msg'] = f'Excel Processing error: {err}'
            return result

        return result
    def main(self,input):
        result = self.check_input(input)
        if result['status'] == 'success':
            self.input_data = input
            template_fields = self.choose_template()
            if (template_fields):
                result = self.check_main(template_fields)
                if result['status'] =='success':
                    result['msg'] = 'Excel is Successfully parsed, Saved in Output folder'
            else:
                result['status']='error'
                result['msg']=f'Reading Template file failed:'
                return result
        return result

 